import streamlit as st
from supabase import create_client, Client
import datetime
import pandas as pd
import json
from fpdf import FPDF
import unicodedata
import re
import os
import io
import time
import requests
import openpyxl
from zoneinfo import ZoneInfo
import hmac
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# --- CONFIGURAÇÕES DA PÁGINA & TEMA APROAR (CLARO / AZUL) ---
st.set_page_config(page_title="APROAR - Controle de Presenças", page_icon="👷", layout="wide")

# Paleta principal. Se a identidade visual mudar, basta alterar o azul aqui e no CSS abaixo.
AZUL_APROAR = "#2563EB"
AZUL_APROAR_ESCURO = "#1D4ED8"

st.markdown("""
    <style>
    :root {
        --aproar-blue: #2563EB;
        --aproar-blue-dark: #1D4ED8;
        --aproar-blue-soft: #EFF6FF;
        --aproar-bg: #FFFFFF;
        --aproar-sidebar: #0F172A;
        --aproar-text: #0F172A;
        --aproar-muted: #64748B;
        --aproar-border: #E2E8F0;
    }

    html, body, [data-testid="stAppViewContainer"], .stApp,
    [data-testid="stMain"], .main {
        background-color: var(--aproar-bg) !important;
        color: var(--aproar-text) !important;
    }

    [data-testid="stHeader"] {
        background: rgba(255,255,255,0.96) !important;
        border-bottom: 1px solid #F1F5F9 !important;
    }

    h1, h2, h3, h4, h5, h6, p, label,
    .stMarkdown, .stText {
        color: var(--aproar-text) !important;
        font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
    }

    /* Não sobrescreve a fonte dos ícones internos do Streamlit.
       Isso evita aparecer texto como _arrow_right no lugar das setas. */
    .material-symbols-rounded, .material-symbols-outlined,
    [data-testid="stIconMaterial"] {
        font-family: "Material Symbols Rounded", "Material Symbols Outlined" !important;
    }

    small, .stCaption, [data-testid="stCaptionContainer"],
    [data-testid="stCaptionContainer"] p {
        color: var(--aproar-muted) !important;
    }

    section[data-testid="stSidebar"] {
        background-color: #0F172A !important;
        border-right: 1px solid #1E293B !important;
        width: 240px !important;
        min-width: 240px !important;
        padding-top: 15px;
    }
    section[data-testid="stSidebar"] > div:first-child {
        padding-left: 15px;
        padding-right: 15px;
    }
    section[data-testid="stSidebar"] p,
    section[data-testid="stSidebar"] label,
    section[data-testid="stSidebar"] .stMarkdown,
    section[data-testid="stSidebar"] [data-testid="stCaptionContainer"],
    section[data-testid="stSidebar"] [data-testid="stCaptionContainer"] p {
        color: #E2E8F0 !important;
    }
    section[data-testid="stSidebar"] hr {
        border-color: #334155 !important;
    }
    section[data-testid="stSidebar"] [data-testid="stImage"] {
        margin-top: 4px;
        margin-bottom: 2px;
    }
    .aproar-sidebar-section {
        color: #94A3B8 !important;
        font-size: 10px !important;
        line-height: 1.2 !important;
        letter-spacing: 1.35px !important;
        font-weight: 800 !important;
        margin: 17px 2px 7px 2px !important;
        text-transform: uppercase;
    }

    /* Campos de formulário */
    div[data-baseweb="select"] > div,
    div[data-baseweb="base-input"] > div,
    div[data-baseweb="input"] > div,
    [data-baseweb="textarea"] > div,
    input, textarea, div[role="combobox"] {
        background-color: #FFFFFF !important;
        color: var(--aproar-text) !important;
        border-color: #CBD5E1 !important;
        border-radius: 8px !important;
    }
    input::placeholder, textarea::placeholder {
        color: #94A3B8 !important;
    }
    ul[data-baseweb="menu"], div[data-baseweb="popover"] {
        background-color: #FFFFFF !important;
        color: var(--aproar-text) !important;
    }
    li[role="option"] {
        background-color: #FFFFFF !important;
        color: var(--aproar-text) !important;
    }
    li[role="option"]:hover, li[role="option"][aria-selected="true"] {
        background-color: var(--aproar-blue-soft) !important;
        color: var(--aproar-blue-dark) !important;
    }

    /* Tags do multiselect */
    div[data-baseweb="tag"] {
        background-color: var(--aproar-blue) !important;
        color: #FFFFFF !important;
    }
    div[data-baseweb="tag"] * { color: #FFFFFF !important; }

    /* Botões */
    .stButton > button,
    .stDownloadButton > button,
    [data-testid="stFormSubmitButton"] > button,
    [data-testid="stFileUploader"] button {
        background: var(--aproar-blue) !important;
        color: #FFFFFF !important;
        border: 1px solid var(--aproar-blue) !important;
        border-radius: 8px !important;
        font-weight: 600 !important;
        box-shadow: none !important;
        transition: all 0.18s ease;
    }
    .stButton > button *, .stDownloadButton > button *,
    [data-testid="stFormSubmitButton"] > button *,
    [data-testid="stFileUploader"] button * {
        color: #FFFFFF !important;
    }
    .stButton > button:hover,
    .stDownloadButton > button:hover,
    [data-testid="stFormSubmitButton"] > button:hover,
    [data-testid="stFileUploader"] button:hover {
        background: var(--aproar-blue-dark) !important;
        border-color: var(--aproar-blue-dark) !important;
        transform: translateY(-1px);
    }

    /* Navegação lateral: azul Aproar sobre fundo escuro */
    section[data-testid="stSidebar"] .stButton > button {
        min-height: 40px !important;
        margin-bottom: 5px !important;
        justify-content: flex-start !important;
        padding-left: 14px !important;
        background: #2563EB !important;
        border-color: #2563EB !important;
        color: #FFFFFF !important;
    }
    section[data-testid="stSidebar"] .stButton > button:hover {
        background: #1D4ED8 !important;
        border-color: #1D4ED8 !important;
    }
    section[data-testid="stSidebar"] .stButton > button * {
        color: #FFFFFF !important;
    }

    /* Evita rolagem horizontal criada por componentes largos no modo wide */
    html, body, [data-testid="stAppViewContainer"], [data-testid="stMain"] {
        overflow-x: hidden !important;
    }
    main .block-container {
        max-width: 100% !important;
        padding-left: 2rem !important;
        padding-right: 2rem !important;
    }
    @media (max-width: 900px) {
        main .block-container {
            padding-left: 1rem !important;
            padding-right: 1rem !important;
        }
    }

    /* Containers e métricas */
    div[data-testid="stVerticalBlock"] > div[style*="border"],
    [data-testid="stMetric"],
    [data-testid="stExpander"] {
        background-color: #FFFFFF !important;
        border-color: var(--aproar-border) !important;
        border-radius: 12px !important;
        box-shadow: 0 1px 2px rgba(15, 23, 42, 0.04) !important;
    }
    [data-testid="stMetricLabel"] *, [data-testid="stMetricValue"] * {
        color: var(--aproar-text) !important;
    }

    /* Abas */
    [data-baseweb="tab-list"] {
        gap: 4px;
        border-bottom: 1px solid var(--aproar-border);
    }
    [data-baseweb="tab"] {
        color: #475569 !important;
        background: transparent !important;
    }
    [data-baseweb="tab"][aria-selected="true"] {
        color: var(--aproar-blue) !important;
        font-weight: 700 !important;
    }
    [data-baseweb="tab-highlight"] {
        background-color: var(--aproar-blue) !important;
    }

    /* Tabelas / editor */
    [data-testid="stDataFrame"], [data-testid="stDataEditor"] {
        border: 1px solid var(--aproar-border) !important;
        border-radius: 10px !important;
        overflow: hidden;
    }

    /* Upload */
    [data-testid="stFileUploaderDropzone"] {
        background: #F8FAFC !important;
        border-color: #CBD5E1 !important;
    }
    [data-testid="stFileUploaderDropzone"] * {
        color: var(--aproar-text) !important;
    }

    /* Alertas continuam coloridos, mas com texto legível */
    [data-testid="stAlert"] p, [data-testid="stAlert"] span {
        color: inherit !important;
    }

    hr { border-color: var(--aproar-border) !important; }
    </style>
""", unsafe_allow_html=True)

# --- MESES EM PORTUGUÊS ---
MESES_PT = {
    1: "JANEIRO", 2: "FEVEREIRO", 3: "MARÇO", 4: "ABRIL",
    5: "MAIO", 6: "JUNHO", 7: "JULHO", 8: "AGOSTO",
    9: "SETEMBRO", 10: "OUTUBRO", 11: "NOVEMBRO", 12: "DEZEMBRO"
}

# --- BANCO DE DADOS: NEON (PREFERENCIAL) OU SUPABASE (FALLBACK) ---
#
# Se DATABASE_URL existir nos Secrets, o sistema usa PostgreSQL/Neon.
# Caso contrário, mantém compatibilidade com o Supabase antigo.
#
# O adaptador abaixo imita a pequena parte da API supabase.table(...)
# que este sistema utiliza. Assim o restante do app não precisa ser reescrito.

class _DBResponse:
    def __init__(self, data=None):
        self.data = data if data is not None else []


def _identificador_sql(nome):
    """Valida e protege nomes de tabela/coluna usados internamente pelo app."""
    nome = str(nome or "").strip()
    if not re.fullmatch(r"[A-Za-z_][A-Za-z0-9_]*", nome):
        raise ValueError(f"Identificador SQL inválido: {nome}")
    return f'"{nome}"'


class _PostgresQuery:
    def __init__(self, db, tabela):
        self.db = db
        self.tabela = tabela
        self.operacao = None
        self.colunas = "*"
        self.payload = None
        self.filtros = []
        self.limite = None

    def select(self, colunas="*"):
        self.operacao = "select"
        self.colunas = colunas or "*"
        return self

    def insert(self, payload):
        self.operacao = "insert"
        self.payload = payload
        return self

    def update(self, payload):
        self.operacao = "update"
        self.payload = payload
        return self

    def delete(self):
        self.operacao = "delete"
        return self

    def eq(self, coluna, valor):
        self.filtros.append(("eq", coluna, valor))
        return self

    def gte(self, coluna, valor):
        self.filtros.append(("gte", coluna, valor))
        return self

    def lte(self, coluna, valor):
        self.filtros.append(("lte", coluna, valor))
        return self

    def in_(self, coluna, valores):
        self.filtros.append(("in", coluna, list(valores or [])))
        return self

    def limit(self, quantidade):
        self.limite = int(quantidade)
        return self

    def _where(self):
        partes = []
        params = []

        for operador, coluna, valor in self.filtros:
            col = _identificador_sql(coluna)

            if operador == "eq":
                if valor is None:
                    partes.append(f"{col} IS NULL")
                else:
                    partes.append(f"{col} = %s")
                    params.append(valor)

            elif operador == "gte":
                partes.append(f"{col} >= %s")
                params.append(valor)

            elif operador == "lte":
                partes.append(f"{col} <= %s")
                params.append(valor)

            elif operador == "in":
                valores = list(valor or [])
                if not valores:
                    partes.append("FALSE")
                else:
                    placeholders = ", ".join(["%s"] * len(valores))
                    partes.append(f"{col} IN ({placeholders})")
                    params.extend(valores)

        sql = (" WHERE " + " AND ".join(partes)) if partes else ""
        return sql, params

    def _colunas_select(self):
        if str(self.colunas).strip() == "*":
            return "*"
        nomes = [c.strip() for c in str(self.colunas).split(",") if c.strip()]
        return ", ".join(_identificador_sql(c) for c in nomes)

    def execute(self):
        tabela = _identificador_sql(self.tabela)
        where_sql, where_params = self._where()

        with self.db._connect() as conn:
            with conn.cursor() as cur:
                if self.operacao == "select":
                    sql = f"SELECT {self._colunas_select()} FROM {tabela}{where_sql}"
                    params = list(where_params)
                    if self.limite is not None:
                        sql += " LIMIT %s"
                        params.append(self.limite)

                    cur.execute(sql, params)
                    rows = cur.fetchall()
                    return _DBResponse([dict(r) for r in rows])

                if self.operacao == "insert":
                    registros = self.payload if isinstance(self.payload, list) else [self.payload]
                    registros = [r for r in registros if isinstance(r, dict) and r]
                    if not registros:
                        return _DBResponse([])

                    resultado = []
                    for registro in registros:
                        colunas = list(registro.keys())
                        cols_sql = ", ".join(_identificador_sql(c) for c in colunas)
                        placeholders = ", ".join(["%s"] * len(colunas))
                        valores = [registro[c] for c in colunas]

                        cur.execute(
                            f"INSERT INTO {tabela} ({cols_sql}) "
                            f"VALUES ({placeholders}) RETURNING *",
                            valores,
                        )
                        row = cur.fetchone()
                        if row:
                            resultado.append(dict(row))

                    conn.commit()
                    return _DBResponse(resultado)

                if self.operacao == "update":
                    payload = dict(self.payload or {})
                    if not payload:
                        return _DBResponse([])

                    sets = []
                    params = []
                    for coluna, valor in payload.items():
                        sets.append(f"{_identificador_sql(coluna)} = %s")
                        params.append(valor)

                    sql = (
                        f"UPDATE {tabela} SET {', '.join(sets)}"
                        f"{where_sql} RETURNING *"
                    )
                    params.extend(where_params)
                    cur.execute(sql, params)
                    rows = cur.fetchall()
                    conn.commit()
                    return _DBResponse([dict(r) for r in rows])

                if self.operacao == "delete":
                    sql = f"DELETE FROM {tabela}{where_sql} RETURNING *"
                    cur.execute(sql, where_params)
                    rows = cur.fetchall()
                    conn.commit()
                    return _DBResponse([dict(r) for r in rows])

                raise RuntimeError("Nenhuma operação de banco foi definida.")


class _PostgresCompat:
    def __init__(self, database_url):
        self.database_url = str(database_url).strip()

    def _connect(self):
        try:
            import psycopg
            from psycopg.rows import dict_row
        except ImportError:
            raise RuntimeError(
                "O pacote psycopg não está instalado. "
                "Adicione psycopg[binary]>=3.2 ao requirements.txt."
            )

        return psycopg.connect(
            self.database_url,
            row_factory=dict_row,
            connect_timeout=12,
        )

    def table(self, tabela):
        if tabela not in {"obras", "colaboradores", "convocacoes"}:
            raise ValueError(f"Tabela não autorizada no adaptador: {tabela}")
        return _PostgresQuery(self, tabela)


def _secret_opcional(nome):
    try:
        valor = st.secrets.get(nome, "")
        return str(valor).strip() if valor else ""
    except Exception:
        return ""


DATABASE_URL = _secret_opcional("DATABASE_URL")
DB_BACKEND = "NEON" if DATABASE_URL else "SUPABASE"


@st.cache_resource
def init_connection():
    if DATABASE_URL:
        return _PostgresCompat(DATABASE_URL)

    url = _secret_opcional("SUPABASE_URL")
    key = _secret_opcional("SUPABASE_KEY") or _secret_opcional("SUPABASE_ANON_KEY")
    if not url or not key:
        try:
            bloco_supabase = st.secrets.get("supabase", {})
            url = url or str(bloco_supabase.get("url", "") or bloco_supabase.get("SUPABASE_URL", "")).strip()
            key = key or str(bloco_supabase.get("key", "") or bloco_supabase.get("anon_key", "") or bloco_supabase.get("SUPABASE_KEY", "")).strip()
        except Exception:
            pass
    if not url or not key:
        raise RuntimeError("Configure DATABASE_URL (Neon) ou SUPABASE_URL + SUPABASE_KEY/SUPABASE_ANON_KEY nos Secrets.")
    return create_client(url, key)


try:
    # Mantemos o nome "supabase" por compatibilidade com o restante do app.
    # Quando DATABASE_URL existe, este objeto na verdade aponta para o Neon/PostgreSQL.
    supabase = init_connection()
except Exception as e:
    st.error(f"Erro ao iniciar o banco de dados: {e}")
    st.stop()


# --- LIMPEZA SELETIVA DE CACHE ---
def limpar_cache_operacional():
    """
    Atualiza apenas os caches do banco local.
    Não limpa o cache do Trello: isso evita uma nova chamada desnecessária
    ao quadro público após cada inclusão, exclusão ou sincronização.
    """
    for nome_funcao in ("buscar_obras", "buscar_colaboradores"):
        funcao = globals().get(nome_funcao)
        if funcao is not None and hasattr(funcao, "clear"):
            try:
                funcao.clear()
            except Exception:
                pass


# --- FUNÇÕES DE LIMPEZA E PADRONIZAÇÃO ---
def identificar_unidade(nome_card):
    if not nome_card: return "GERAL"
    texto = unicodedata.normalize('NFKD', str(nome_card)).encode('ASCII', 'ignore').decode('utf-8').upper()
    
    if "APRL005" in texto or "MARACANAU" in texto: return "MARACANAÚ"
    if "SEBRAE" in texto: return "SEBRAE"
    if "UNIFOR" in texto: return "UNIFOR"
    if "IDALYA" in texto or "MATHEUS" in texto: return "IDALYA E MATHEUS"
    if "COLISEU" in texto: return "COLISEU"
    if "BARRA" in texto: return "BARRA DO CEARÁ"
    if "MUSEU" in texto: return "MUSEU"
    if "HORIZONTE" in texto: return "HORIZONTE"
    if "ESCRITORIO" in texto: return "ESCRITÓRIO"
    if "CASA DA INDUSTRIA" in texto or "FIEC" in texto or " DR " in texto or "| SESI DR |" in texto or "| SESI DR" in texto: return "FIEC"
    if "CENTRO" in texto: return "CENTRO"
    
    partes = str(nome_card).split('|')
    if len(partes) >= 2:
        return partes[1].strip().upper()
    return "GERAL"

def limpar_funcao(texto):
    if not texto or str(texto).upper() == 'NAN': return "INDEFINIDA"
    texto_limpo = str(texto).upper().strip()
    texto_limpo = re.sub(r'^\d+\s*-\s*', '', texto_limpo)
    texto_limpo = unicodedata.normalize('NFKD', texto_limpo).encode('ASCII', 'ignore').decode('utf-8')
    return texto_limpo

def normalizar(texto):
    if not texto: return ""
    return ''.join(c for c in unicodedata.normalize('NFD', str(texto)) if unicodedata.category(c) != 'Mn').upper().strip()

def get_cor_funcao(funcao):
    cores = ["🟥", "🟧", "🟨", "🟩", "🟦", "🟪", "🟫", "⬛"]
    hash_num = sum(ord(c) for c in str(funcao))
    return cores[hash_num % len(cores)]

VALOR_DIARIA_PROFISSIONAL = 241.74
VALOR_DIARIA_AJUDANTE = 182.34

def inferir_tipo_colaborador(funcao):
    """Classifica cadastros antigos quando a diária ainda não está nos novos valores fixos."""
    f = normalizar(funcao or "")
    termos_ajudante = ["AJUDANTE", "AUXILIAR", "AUX.", "AUX ", "SERVENTE"]
    return "Ajudante" if any(t in f for t in termos_ajudante) else "Profissional"

def valor_diaria_por_tipo(tipo):
    return VALOR_DIARIA_AJUDANTE if normalizar(tipo) == "AJUDANTE" else VALOR_DIARIA_PROFISSIONAL

def obter_valor_diaria_colaborador(colab):
    """Aplica R$ 241,74 para profissional e R$ 182,34 para ajudante em todo o sistema."""
    colab = colab or {}
    try:
        valor_cadastrado = float(colab.get("valor_diaria") or 0.0)
    except Exception:
        valor_cadastrado = 0.0

    # Novos cadastros já persistem exatamente um dos dois valores oficiais.
    if abs(valor_cadastrado - VALOR_DIARIA_PROFISSIONAL) < 0.01:
        return VALOR_DIARIA_PROFISSIONAL
    if abs(valor_cadastrado - VALOR_DIARIA_AJUDANTE) < 0.01:
        return VALOR_DIARIA_AJUDANTE

    # Compatibilidade com cadastros antigos (ex.: diária antiga de R$ 240,00).
    return valor_diaria_por_tipo(inferir_tipo_colaborador(colab.get("funcao", "")))

def calcular_diaria_proporcional(status, valor_diaria_base):
    diaria = float(valor_diaria_base or VALOR_DIARIA_PROFISSIONAL)
    if status in ["Presente (Integral)", "Presente", "Extra"]:
        return diaria
    elif status in ["Presente (Só Manhã)", "Presente (Só Tarde)", "Saída Antecipada"]:
        return diaria / 2.0
    return 0.0

def formatar_reais(valor):
    return f"R$ {float(valor):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')

def to_latin(texto):
    if not texto: return ""
    return str(texto).encode('latin-1', 'replace').decode('latin-1')

def proximo_dia_util(data_base=None):
    """
    Retorna o dia seguinte à data informada.
    Convocações podem ocorrer em qualquer dia da semana,
    inclusive sábado, domingo e feriado.

    O nome da função foi mantido para compatibilidade
    com o restante do sistema.
    """
    data_ref = data_base or datetime.date.today()
    return data_ref + datetime.timedelta(days=1)

NOME_OBRA_PLACEHOLDER = "A DEFINIR NO APONTAMENTO"

def nome_obra_placeholder_unidade(unidade):
    """
    Gera um nome técnico único por Unidade.
    Isso evita conflito com o índice UNIQUE de nome da tabela obras no Neon.
    """
    unidade_limpa = " ".join(str(unidade or "").strip().split()).upper()
    return f"{NOME_OBRA_PLACEHOLDER} - {unidade_limpa}"


def eh_obra_placeholder(obra):
    """
    Identifica tanto o placeholder legado:
        A DEFINIR NO APONTAMENTO
    quanto os novos placeholders por Unidade:
        A DEFINIR NO APONTAMENTO - UNIFOR
        A DEFINIR NO APONTAMENTO - FIEC
        etc.
    """
    if not obra:
        return False

    nome_norm = normalizar(obra.get("nome", ""))
    prefixo_norm = normalizar(NOME_OBRA_PLACEHOLDER)

    return nome_norm == prefixo_norm or nome_norm.startswith(prefixo_norm + " - ")


def obter_obra_placeholder_unidade(unidade):
    """
    Retorna/cria a obra técnica da Unidade para convocações ainda sem
    Obra/Serviço definida.

    Cada Unidade recebe um nome técnico diferente no banco para não violar
    a restrição UNIQUE de obras.nome.
    """
    unidade_limpa = " ".join(str(unidade or "").strip().split()).upper()
    if not unidade_limpa:
        return None

    try:
        # 1) Se já existe qualquer placeholder nessa Unidade, reutiliza.
        existentes = (
            supabase.table("obras")
            .select("*")
            .eq("unidade", unidade_limpa)
            .execute().data or []
        )

        for obra in existentes:
            if eh_obra_placeholder(obra):
                return obra.get("id")

        # 2) Cria um placeholder exclusivo desta Unidade.
        nome_placeholder = nome_obra_placeholder_unidade(unidade_limpa)

        criado = (
            supabase.table("obras")
            .insert({
                "unidade": unidade_limpa,
                "nome": nome_placeholder
            })
            .execute().data or []
        )

        if criado:
            limpar_cache_operacional()
            return criado[0].get("id")

    except Exception as e:
        # Guarda o diagnóstico para o administrativo, mas não derruba o app.
        try:
            st.session_state["erro_placeholder_unidade"] = (
                f"{type(e).__name__}: {str(e)[:300]}"
            )
        except Exception:
            pass

        # 3) Última confirmação: o INSERT pode ter sido concluído e apenas a
        # resposta ter falhado. Consulta novamente antes de desistir.
        try:
            existentes = (
                supabase.table("obras")
                .select("*")
                .eq("unidade", unidade_limpa)
                .execute().data or []
            )
            for obra in existentes:
                if eh_obra_placeholder(obra):
                    return obra.get("id")
        except Exception:
            pass

    return None

def obras_reais_da_unidade(unidade):
    """Lista somente Obras/Serviços reais, ocultando o registro temporário."""
    return [o for o in obras if o.get("unidade") == unidade and not eh_obra_placeholder(o)]

OBS_META_MARKER = " ||APROAR_META|| "
try:
    TZ_APROAR = ZoneInfo("America/Fortaleza")
except Exception:
    TZ_APROAR = None


def agora_aproar():
    """Horário operacional da Aproar (Fortaleza), inclusive no Streamlit Cloud."""
    if TZ_APROAR is not None:
        return datetime.datetime.now(TZ_APROAR)
    return datetime.datetime.now()


def separar_observacao_metadata(observacao):
    texto = str(observacao or "")
    if OBS_META_MARKER not in texto:
        return texto.strip(), {}
    base, bruto = texto.split(OBS_META_MARKER, 1)
    try:
        meta = json.loads(bruto.strip()) if bruto.strip() else {}
        if not isinstance(meta, dict):
            meta = {}
    except Exception:
        meta = {}
    return base.strip(), meta


def obter_metadata_operacional(observacao):
    return separar_observacao_metadata(observacao)[1]


def decompor_observacao_operacional(observacao):
    """Lê turno/observação livre escondendo os metadados internos de auditoria."""
    texto, _ = separar_observacao_metadata(observacao)
    turno = "Integral"

    m_turno = re.search(r"Turno:\s*(Integral|Manhã|Tarde|Noite)", texto, flags=re.IGNORECASE)
    if m_turno:
        turno_encontrado = m_turno.group(1).lower()
        mapa_turnos = {"integral": "Integral", "manhã": "Manhã", "tarde": "Tarde", "noite": "Noite"}
        turno = mapa_turnos.get(turno_encontrado, "Integral")

    livre = re.sub(r"Turno:\s*(Integral|Manhã|Tarde|Noite)\s*(?:\|\s*)?", "", texto, flags=re.IGNORECASE)
    livre = re.sub(r"HE:\s*[0-9]+(?:[\.,][0-9]+)?\s*h\s*(?:\|\s*)?", "", livre, flags=re.IGNORECASE)
    livre = re.sub(r"^Obs:\s*", "", livre, flags=re.IGNORECASE).strip(" |")
    return turno, livre


def montar_observacao_operacional(turno, observacao_livre="", metadata=None):
    partes = [f"Turno: {turno}"]
    if str(observacao_livre or "").strip():
        partes.append(f"Obs: {str(observacao_livre).strip()}")
    texto = " | ".join(partes)
    if metadata:
        try:
            meta_limpa = {k: v for k, v in dict(metadata).items() if v not in (None, "", [], {})}
            if meta_limpa:
                texto += OBS_META_MARKER + json.dumps(meta_limpa, ensure_ascii=False, separators=(",", ":"))
        except Exception:
            pass
    return texto


def atualizar_metadata_observacao(observacao, **alteracoes):
    turno, livre = decompor_observacao_operacional(observacao)
    meta = obter_metadata_operacional(observacao)
    for chave, valor in alteracoes.items():
        if valor is None:
            meta.pop(chave, None)
        else:
            meta[chave] = valor
    return montar_observacao_operacional(turno, livre, meta)


def normalizar_status_operacional(status):
    """Compatibilidade: registros antigos com status Extra passam a ser presença integral."""
    return "Presente (Integral)" if str(status or "") == "Extra" else str(status or "Presente (Integral)")


def status_eh_presenca(status):
    s = normalizar_status_operacional(status)
    return s in ["Presente (Integral)", "Presente (Só Manhã)", "Presente (Só Tarde)", "Saída Antecipada", "Presente"]


def _normalizar_servicos_adicionais(meta):
    """Compatibilidade entre o modelo antigo (lista de nomes) e o novo (serviço + período)."""
    meta = meta or {}
    saida = []
    vistos = set()

    for item in meta.get("servicos_adicionais", []) or []:
        if isinstance(item, dict):
            nome = str(item.get("servico") or "").strip()
            periodo = str(item.get("periodo") or "").strip() or "Não informado"
        else:
            nome = str(item or "").strip()
            periodo = "Não informado"
        chave = normalizar(nome)
        if nome and chave not in vistos:
            saida.append({"servico": nome, "periodo": periodo})
            vistos.add(chave)

    # Registros das versões anteriores continuam válidos.
    for nome in meta.get("servicos_extras", []) or []:
        nome = str(nome or "").strip()
        chave = normalizar(nome)
        if nome and chave not in vistos:
            saida.append({"servico": nome, "periodo": "Não informado"})
            vistos.add(chave)
    return saida


def descricao_servicos_convocacao(convocacao, obra_primaria=None):
    obra_primaria = obra_primaria or dict_obras.get(convocacao.get("obra_id"), {})
    meta = obter_metadata_operacional(convocacao.get("observacao") or "")
    partes = []

    nome_principal = str(obra_primaria.get("nome") or "").strip()
    periodo_principal = str(meta.get("periodo_servico_principal") or "").strip()
    if nome_principal and not eh_obra_placeholder(obra_primaria):
        partes.append(f"{nome_principal} ({periodo_principal})" if periodo_principal else nome_principal)

    for item in _normalizar_servicos_adicionais(meta):
        nome = item["servico"]
        periodo = item.get("periodo") or ""
        rotulo = f"{nome} ({periodo})" if periodo and periodo != "Não informado" else nome
        if normalizar(nome) != normalizar(nome_principal):
            partes.append(rotulo)

    return " + ".join(partes) if partes else NOME_OBRA_PLACEHOLDER


def registrar_metadata_apontamento(
    convocacao,
    data_servico,
    servicos_extras=None,
    apontado_por=None,
    periodo_principal=None,
    servicos_adicionais=None,
):
    meta = obter_metadata_operacional(convocacao.get("observacao") or "")
    agora = agora_aproar()
    if not meta.get("apontado_em"):
        meta["apontado_em"] = agora.isoformat()
    meta["ultimo_apontamento_em"] = agora.isoformat()
    meta["apontamento_atrasado"] = bool(agora.date() > data_servico)
    if apontado_por:
        meta["apontado_por"] = str(apontado_por)
    if periodo_principal:
        meta["periodo_servico_principal"] = str(periodo_principal)

    adicionais = []
    for item in servicos_adicionais or []:
        if not isinstance(item, dict):
            continue
        nome = str(item.get("servico") or "").strip()
        periodo = str(item.get("periodo") or "Não informado").strip()
        if nome:
            adicionais.append({"servico": nome, "periodo": periodo})

    # Compatibilidade com chamadas antigas.
    if not adicionais and servicos_extras:
        adicionais = [{"servico": str(nome), "periodo": "Não informado"} for nome in servicos_extras if str(nome).strip()]

    meta["servicos_adicionais"] = adicionais
    meta["servicos_extras"] = [x["servico"] for x in adicionais]
    return meta


def rotulo_atraso_apontamento(convocacao):
    meta = obter_metadata_operacional(convocacao.get("observacao") or "")
    if meta.get("apontamento_atrasado"):
        return "🟧 APONTAMENTO RETROATIVO"
    return ""

def formatar_nome_whatsapp(nome):
    """Deixa nomes em formato legível para a mensagem, preservando partículas comuns."""
    nome_fmt = " ".join(str(nome or "").strip().split()).title()
    if not nome_fmt:
        return "Colaborador não identificado"
    minusculas = {"Da", "Das", "De", "Do", "Dos", "E"}
    partes = nome_fmt.split()
    return " ".join(p.lower() if i > 0 and p in minusculas else p for i, p in enumerate(partes))


def formatar_unidade_whatsapp(unidade):
    """Formata a Unidade para o cabeçalho da mensagem do WhatsApp."""
    texto = " ".join(str(unidade or "").strip().split())
    if not texto:
        return "Unidade não identificada"
    siglas = {"FIEC", "SEBRAE", "UNIFOR"}
    if normalizar(texto) in siglas:
        return normalizar(texto)
    titulo = texto.title()
    minusculas = {"Da", "Das", "De", "Do", "Dos", "E"}
    partes = titulo.split()
    return " ".join(p.lower() if i > 0 and p in minusculas else p for i, p in enumerate(partes))


def rotulo_data_whatsapp(data_alvo):
    hoje = datetime.date.today()
    if data_alvo == hoje:
        return f"hoje {data_alvo.strftime('%d/%m')}"
    if data_alvo == hoje + datetime.timedelta(days=1):
        return f"amanhã {data_alvo.strftime('%d/%m')}"
    return f"o dia {data_alvo.strftime('%d/%m')}"


def organizar_convocacoes_whatsapp(convocacoes, mostrar_funcao=False):
    """Agrupa convocações por Unidade e Turno para montar a mensagem pronta para copiar."""
    ordem_turnos = ["Integral", "Manhã", "Tarde", "Noite"]
    agrupado = {}

    for conv in convocacoes or []:
        obra = dict_obras.get(conv.get("obra_id"), {})
        unidade = obra.get("unidade") or "NÃO IDENTIFICADA"
        colab = dict_colaboradores.get(conv.get("colaborador_id"), {})
        nome = formatar_nome_whatsapp(colab.get("nome", ""))
        funcao = str(colab.get("funcao", "") or "").strip()
        turno, _ = decompor_observacao_operacional(conv.get("observacao", ""))

        if mostrar_funcao and funcao:
            funcao_fmt = funcao.replace("AVULSO - ", "").strip().title()
            nome = f"{nome} ({funcao_fmt})"

        agrupado.setdefault(unidade, {}).setdefault(turno, [])
        if nome not in agrupado[unidade][turno]:
            agrupado[unidade][turno].append(nome)

    # Ordena colaboradores alfabeticamente e turnos na sequência operacional.
    saida = {}
    for unidade in sorted(agrupado.keys(), key=lambda x: normalizar(x)):
        saida[unidade] = {}
        for turno in ordem_turnos:
            nomes = agrupado[unidade].get(turno, [])
            if nomes:
                saida[unidade][turno] = sorted(nomes, key=lambda x: normalizar(x))
        # Compatibilidade para algum turno antigo/não previsto.
        for turno, nomes in agrupado[unidade].items():
            if turno not in saida[unidade] and nomes:
                saida[unidade][turno] = sorted(nomes, key=lambda x: normalizar(x))
    return saida


def montar_mensagem_whatsapp(data_alvo, convocacoes, mostrar_funcao=False, aviso_pendentes=False, somente_unidade=None):
    agrupado = organizar_convocacoes_whatsapp(convocacoes, mostrar_funcao=mostrar_funcao)
    if somente_unidade is not None:
        agrupado = {somente_unidade: agrupado.get(somente_unidade, {})} if somente_unidade in agrupado else {}

    linhas = [f"Segue divisão de Equipes para {rotulo_data_whatsapp(data_alvo)}", ""]

    for unidade, turnos in agrupado.items():
        linhas.append(f"*{formatar_unidade_whatsapp(unidade)}*")
        linhas.append("")

        turnos_com_pessoas = [(turno, nomes) for turno, nomes in turnos.items() if nomes]
        exibir_turnos = len(turnos_com_pessoas) > 1 or any(turno != "Integral" for turno, _ in turnos_com_pessoas)

        contador = 1
        for turno, nomes in turnos_com_pessoas:
            if exibir_turnos:
                linhas.append(f"_{turno}_")
            for nome in nomes:
                linhas.append(f"{contador}. {nome}")
                contador += 1
            if exibir_turnos:
                linhas.append("")
        linhas.append("")

    if aviso_pendentes:
        if agrupado:
            linhas.append("As demais demandas serão enviadas pelos respectivos responsáveis.")
        else:
            linhas.append("As demandas serão enviadas pelos respectivos responsáveis.")

    # Remove excesso de linhas vazias no fim sem mexer na separação interna.
    while linhas and not str(linhas[-1]).strip():
        linhas.pop()
    return "\n".join(linhas)

def criar_ou_obter_colaborador_manual(nome, tipo, funcao_livre="", avulso=False):
    """Cria um colaborador digitado pelo engenheiro sem exigir novas colunas no Supabase."""
    nome_limpo = " ".join(str(nome or "").strip().split())
    if not nome_limpo:
        return None, None, "Informe o nome do colaborador."

    try:
        atuais = supabase.table("colaboradores").select("*").execute().data or []
        existente = next((c for c in atuais if normalizar(c.get("nome", "")) == normalizar(nome_limpo)), None)
        if existente:
            return existente.get("id"), existente, "Cadastro existente localizado e reutilizado."

        funcao_texto = str(funcao_livre or "").strip()
        if avulso:
            funcao_salva = f"AVULSO - {funcao_texto}" if funcao_texto else f"AVULSO - {str(tipo).upper()}"
        else:
            funcao_salva = funcao_texto if funcao_texto else str(tipo).upper()

        valor = valor_diaria_por_tipo(tipo)
        criado = supabase.table("colaboradores").insert({
            "nome": nome_limpo.upper(),
            "funcao": limpar_funcao(funcao_salva),
            "valor_diaria": valor
        }).execute().data or []

        if criado:
            limpar_cache_operacional()
            return criado[0].get("id"), criado[0], "Novo colaborador cadastrado."
        return None, None, "O cadastro não retornou um identificador."
    except Exception:
        return None, None, "Não foi possível cadastrar o nome informado."

def gerar_excel_colaboradores(lista_colaboradores):
    """Gera uma planilha Excel com a base atual de colaboradores do Supabase."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Colaboradores"

    # Título e metadados
    ws.merge_cells("A1:E1")
    ws["A1"] = "APROAR ENGENHARIA - BASE ATUALIZADA DE COLABORADORES"
    ws["A1"].font = Font(name="Arial", size=13, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    ws.merge_cells("A2:E2")
    ws["A2"] = f"Gerado em: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')} | Total de colaboradores: {len(lista_colaboradores)}"
    ws["A2"].font = Font(name="Arial", size=9, italic=True, color="64748B")
    ws["A2"].alignment = Alignment(horizontal="left")

    headers = ["Nome", "Função", "Categoria", "Valor da Diária (R$)", "Avulso"]
    linha_header = 4
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=linha_header, column=col_idx, value=header)
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    borda = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1")
    )

    ordenados = sorted(lista_colaboradores, key=lambda c: normalizar(c.get("nome", "")))
    for row_idx, colab in enumerate(ordenados, linha_header + 1):
        funcao = str(colab.get("funcao") or "").strip()
        valor_diaria = obter_valor_diaria_colaborador(colab)
        categoria = "Ajudante" if abs(valor_diaria - VALOR_DIARIA_AJUDANTE) < 0.01 else "Profissional"
        avulso = "SIM" if normalizar(funcao).startswith("AVULSO -") else "NÃO"

        valores = [
            str(colab.get("nome") or "").strip(),
            funcao,
            categoria,
            valor_diaria,
            avulso,
        ]
        for col_idx, valor in enumerate(valores, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.font = Font(name="Arial", size=9)
            cell.border = borda
            cell.alignment = Alignment(vertical="center", horizontal="left" if col_idx in [1, 2] else "center")
            if col_idx == 4:
                cell.number_format = 'R$ #,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")

    fim = linha_header + len(ordenados)
    if fim >= linha_header:
        ws.auto_filter.ref = f"A{linha_header}:E{fim}"
    ws.freeze_panes = "A5"

    larguras = {"A": 38, "B": 30, "C": 16, "D": 22, "E": 12}
    for coluna, largura in larguras.items():
        ws.column_dimensions[coluna].width = largura

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()

def buscar_convocacao_existente(colaborador_id, data_convocacao):
    try:
        return (
            supabase.table("convocacoes")
            .select("*")
            .eq("colaborador_id", colaborador_id)
            .eq("data", data_convocacao.isoformat())
            .execute().data or []
        )
    except Exception:
        return []

def registrar_conflito_convocacao(registro_existente, engenheiro_tentativa):
    """Persiste a tentativa conflitante na própria convocação para aparecer ao Paulo/Admin."""
    try:
        obs_atual = registro_existente.get("observacao") or ""
        meta = obter_metadata_operacional(obs_atual)
        conflitos = list(meta.get("conflitos_convocacao") or [])
        conflitos.append({
            "tentativa_por": str(engenheiro_tentativa or "N/A"),
            "em": agora_aproar().isoformat(),
            "resolvido": False,
        })
        meta["conflitos_convocacao"] = conflitos[-20:]
        turno, livre = decompor_observacao_operacional(obs_atual)
        nova_obs = montar_observacao_operacional(turno, livre, meta)
        supabase.table("convocacoes").update({"observacao": nova_obs}).eq("id", registro_existente.get("id")).execute()
        limpar_cache_operacional()
        return True
    except Exception:
        return False


def resolver_conflitos_convocacao(registro):
    try:
        obs_atual = registro.get("observacao") or ""
        meta = obter_metadata_operacional(obs_atual)
        conflitos = list(meta.get("conflitos_convocacao") or [])
        for conflito in conflitos:
            conflito["resolvido"] = True
            conflito["resolvido_em"] = agora_aproar().isoformat()
        meta["conflitos_convocacao"] = conflitos
        turno, livre = decompor_observacao_operacional(obs_atual)
        supabase.table("convocacoes").update({
            "observacao": montar_observacao_operacional(turno, livre, meta)
        }).eq("id", registro.get("id")).execute()
        limpar_cache_operacional()
        return True
    except Exception:
        return False


def inserir_convocacao_segura(obra_id, colaborador_id, data_convocacao, engenheiro, turno):
    """Bloqueia indisponibilidade/duplicidade e registra conflitos para auditoria."""
    indisp = obter_indisponibilidade_colaborador(colaborador_id, data_convocacao)
    if indisp:
        return False, (
            f"está indisponível ({indisp.get('motivo', 'Indisponível')}) de "
            f"{indisp.get('inicio', '')} a {indisp.get('fim', '')}"
        )

    existente = buscar_convocacao_existente(colaborador_id, data_convocacao)
    if existente:
        reg = existente[0]
        eng_atual = str(reg.get("engenheiro") or "N/A")
        if normalizar(eng_atual) != normalizar(engenheiro):
            registrado = registrar_conflito_convocacao(reg, engenheiro)
            complemento = " O conflito foi registrado para conferência do Paulo." if registrado else ""
            return False, f"já foi convocado(a) por {eng_atual} para esta data.{complemento}"
        return False, f"já estava convocado(a) por você para esta data"

    agora = agora_aproar()
    meta = {
        "convocado_em": agora.isoformat(),
        "convocado_por": str(engenheiro),
        "convocacao_atrasada": bool(
            agora.hour >= 16 and data_convocacao == proximo_dia_util(agora.date())
        ),
    }
    try:
        supabase.table("convocacoes").insert({
            "obra_id": obra_id,
            "colaborador_id": colaborador_id,
            "data": data_convocacao.isoformat(),
            "engenheiro": engenheiro,
            "status": "Presente (Integral)",
            "valor_extra": 0,
            "observacao": montar_observacao_operacional(turno, "", meta)
        }).execute()
        return True, "convocado(a) com sucesso"
    except Exception:
        return False, "não pôde ser convocado(a); verifique os dados e tente novamente"


# --- ACESSO RESILIENTE AO BANCO ---
def _cliente_supabase_para_tentativa(tentativa=0):
    """
    Retorna um cliente novo para retry.
    O nome da função foi preservado para compatibilidade interna.
    """
    if DB_BACKEND == "NEON":
        return _PostgresCompat(DATABASE_URL)

    if tentativa == 0:
        return supabase

    url = _secret_opcional("SUPABASE_URL")
    key = _secret_opcional("SUPABASE_KEY") or _secret_opcional("SUPABASE_ANON_KEY")
    if not url or not key:
        try:
            bloco_supabase = st.secrets.get("supabase", {})
            url = url or str(bloco_supabase.get("url", "") or bloco_supabase.get("SUPABASE_URL", "")).strip()
            key = key or str(bloco_supabase.get("key", "") or bloco_supabase.get("anon_key", "") or bloco_supabase.get("SUPABASE_KEY", "")).strip()
        except Exception:
            pass
    return create_client(url, key)


def _executar_supabase_com_retry(operacao, tentativas=3):
    ultimo_erro = None

    for tentativa in range(tentativas):
        try:
            cliente = _cliente_supabase_para_tentativa(tentativa)
            return operacao(cliente)
        except Exception as e:
            ultimo_erro = e
            try:
                st.cache_resource.clear()
            except Exception:
                pass

            if tentativa < tentativas - 1:
                time.sleep(1.0 * (tentativa + 1))

    raise ultimo_erro


def _listar_obras_resiliente():
    resposta = _executar_supabase_com_retry(
        lambda db: db.table("obras").select("id,nome,unidade").execute(),
        tentativas=3
    )
    return resposta.data or [], DB_BACKEND.lower()


def _obra_existe_no_supabase(nome_obra, tentativas=2):
    resposta = _executar_supabase_com_retry(
        lambda db: (
            db.table("obras")
            .select("id,nome")
            .eq("nome", nome_obra)
            .limit(1)
            .execute()
        ),
        tentativas=tentativas
    )
    return bool(resposta.data or [])


def _inserir_obra_resiliente(nome_obra, unidade, tentativas=3):
    ultimo_erro = None

    for tentativa in range(tentativas):
        try:
            if _obra_existe_no_supabase(nome_obra, tentativas=1):
                return True, "existente"
        except Exception as e:
            ultimo_erro = e

        try:
            cliente = _cliente_supabase_para_tentativa(tentativa)
            (
                cliente.table("obras")
                .insert({
                    "unidade": unidade,
                    "nome": nome_obra
                })
                .execute()
            )
            return True, "inserida"

        except Exception as e:
            ultimo_erro = e

            try:
                if _obra_existe_no_supabase(nome_obra, tentativas=1):
                    return True, "inserida"
            except Exception:
                pass

            if tentativa < tentativas - 1:
                time.sleep(1.2 * (tentativa + 1))

    try:
        st.session_state["banco_ultimo_erro_sync"] = (
            f"{type(ultimo_erro).__name__}: {str(ultimo_erro)[:220]}"
        )
    except Exception:
        pass

    return False, "erro"


def _testar_banco_ativo():
    try:
        resposta = _executar_supabase_com_retry(
            lambda db: db.table("obras").select("id").limit(1).execute(),
            tentativas=2
        )
        return True, f"{DB_BACKEND} OK"
    except Exception as e:
        return False, f"{type(e).__name__}: {str(e)[:220]}"


# --- SINCRONIZAÇÃO COM TRELLO (MÊS VIGENTE OU SELEÇÃO MANUAL) ---
TRELLO_JSON_URL = "https://trello.com/b/TX8hGvmI.json"


def _garantir_tabela_snapshot_trello():
    """
    Cria uma tabela minúscula no Neon para guardar a última leitura válida
    do quadro público. Isso permite o sistema continuar sincronizando mesmo
    se o Trello estiver temporariamente lento/fora do ar.
    """
    if DB_BACKEND != "NEON" or not hasattr(supabase, "_connect"):
        return False

    try:
        with supabase._connect() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    CREATE TABLE IF NOT EXISTS trello_snapshot (
                        snapshot_id INTEGER PRIMARY KEY,
                        listas JSONB NOT NULL,
                        cards JSONB NOT NULL,
                        atualizado_em TIMESTAMPTZ NOT NULL DEFAULT NOW()
                    )
                    """
                )
                conn.commit()
        return True
    except Exception:
        return False


def _salvar_snapshot_trello(listas, cards):
    if not _garantir_tabela_snapshot_trello():
        return

    try:
        with supabase._connect() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    INSERT INTO trello_snapshot
                        (snapshot_id, listas, cards, atualizado_em)
                    VALUES
                        (1, %s::jsonb, %s::jsonb, NOW())
                    ON CONFLICT (snapshot_id)
                    DO UPDATE SET
                        listas = EXCLUDED.listas,
                        cards = EXCLUDED.cards,
                        atualizado_em = NOW()
                    """,
                    (
                        json.dumps(listas, ensure_ascii=False),
                        json.dumps(cards, ensure_ascii=False),
                    )
                )
                conn.commit()
    except Exception:
        pass


def _carregar_snapshot_trello():
    if not _garantir_tabela_snapshot_trello():
        return [], [], None

    try:
        with supabase._connect() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT listas, cards, atualizado_em
                    FROM trello_snapshot
                    WHERE snapshot_id = 1
                    LIMIT 1
                    """
                )
                linha = cur.fetchone()

        if not linha:
            return [], [], None

        if isinstance(linha, dict):
            listas = linha.get("listas") or []
            cards = linha.get("cards") or []
            atualizado_em = linha.get("atualizado_em")
        else:
            listas, cards, atualizado_em = linha

        if isinstance(listas, list) and isinstance(cards, list):
            return listas, cards, atualizado_em

    except Exception:
        pass

    return [], [], None


@st.cache_data(ttl=300, show_spinner=False)
def _baixar_trello_publico():
    """
    Leitura pública do mesmo .json usado pela Torre de Controle.
    Sem API Key, Token ou autenticação.
    Faz uma segunda tentativa somente quando necessário.
    """
    ultimo_erro = None

    for tentativa in range(2):
        try:
            # Timeout separado: conexão rápida, mas tolerância maior para leitura
            # do JSON completo quando o Trello estiver mais lento.
            timeout_leitura = 35 if tentativa == 0 else 65

            resposta = requests.get(
                TRELLO_JSON_URL,
                timeout=(8, timeout_leitura)
            )
            resposta.raise_for_status()

            dados = resposta.json()

            if (
                not isinstance(dados, dict)
                or not isinstance(dados.get("cards"), list)
                or not isinstance(dados.get("lists"), list)
            ):
                raise ValueError(
                    "O JSON público respondeu sem as listas/cards esperados."
                )

            return dados.get("lists", []), dados.get("cards", [])

        except Exception as e:
            ultimo_erro = e

            if tentativa == 0:
                time.sleep(1.2)

    raise ultimo_erro


def obter_listas_trello():
    """
    Ordem de prioridade:
      1. Trello público ao vivo / cache de 5 minutos;
      2. última leitura válida da sessão;
      3. último snapshot persistido no Neon.

    O sistema não fica inutilizável por um timeout momentâneo do Trello.
    """
    try:
        listas, cards = _baixar_trello_publico()

        if isinstance(listas, list) and isinstance(cards, list):
            st.session_state["trello_snapshot_sessao"] = {
                "listas": listas,
                "cards": cards,
            }
            st.session_state["trello_ultimo_erro"] = ""
            st.session_state["trello_fonte"] = "Trello público"
            st.session_state["trello_usando_snapshot"] = False

            # Persistência de contingência; falha silenciosa não afeta o app.
            _salvar_snapshot_trello(listas, cards)

            return listas, cards

    except Exception as e:
        st.session_state["trello_ultimo_erro"] = (
            f"{type(e).__name__}: {str(e)[:300]}"
        )

    # Fallback 1: última leitura desta sessão
    snapshot_sessao = st.session_state.get("trello_snapshot_sessao") or {}
    listas_sessao = snapshot_sessao.get("listas") or []
    cards_sessao = snapshot_sessao.get("cards") or []

    if listas_sessao or cards_sessao:
        st.session_state["trello_fonte"] = "Última leitura válida"
        st.session_state["trello_usando_snapshot"] = True
        return listas_sessao, cards_sessao

    # Fallback 2: snapshot persistente do Neon
    listas_db, cards_db, atualizado_em = _carregar_snapshot_trello()

    if listas_db or cards_db:
        st.session_state["trello_snapshot_sessao"] = {
            "listas": listas_db,
            "cards": cards_db,
        }
        st.session_state["trello_fonte"] = "Última leitura salva"
        st.session_state["trello_usando_snapshot"] = True
        st.session_state["trello_snapshot_data"] = atualizado_em
        return listas_db, cards_db

    st.session_state["trello_fonte"] = ""
    st.session_state["trello_usando_snapshot"] = False
    return [], []


def executar_sincronizacao_trello(id_lista_target=None, id_card_target=None, listas_precarregadas=None, cards_precarregados=None):
    """
    Sincroniza cards/listas do Trello com a tabela 'obras'.

    Sincroniza o Trello público com o banco ativo (Neon/PostgreSQL).
    A leitura do Trello é sempre renovada ao executar uma sincronização.
    """
    try:
        # A tela de Configurações já leu o Trello para montar os dropdowns.
        # Reutilizamos exatamente esse payload no clique do botão, evitando
        # uma segunda requisição que era a causa do ReadTimeout.
        if listas_precarregadas is not None and cards_precarregados is not None:
            lists = list(listas_precarregadas)
            cards = list(cards_precarregados)
        else:
            lists, cards = obter_listas_trello()
        if not lists and not cards:
            detalhe = ""
            try:
                detalhe = st.session_state.get("trello_ultimo_erro", "")
            except Exception:
                pass

            msg = (
                "Não foi possível obter uma leitura válida do quadro ORÇAMENTOS agora. "
                "O restante do sistema continua funcionando normalmente."
            )
            return False, msg

        nome_alvo = ""
        cards_execucao = []

        # Busca manual de um card específico (útil para medições retroativas)
        if id_card_target:
            card_alvo = next((c for c in cards if c.get("id") == id_card_target), None)
            if not card_alvo:
                return False, "Card selecionado não foi encontrado no Trello."

            cards_execucao = [card_alvo]
            nome_alvo = f"Card: {card_alvo.get('name', 'Sem nome')}"

        else:
            id_lista_execucao = id_lista_target
            nome_lista_alvo = ""

            # Sem seleção manual: procura primeiro a medição do mês vigente.
            if not id_lista_execucao:
                hoje = datetime.date.today()
                mes_vigente = MESES_PT.get(hoje.month, "")
                ano_vigente = str(hoje.year)
                termo_busca = f"MEDICAO {mes_vigente} {ano_vigente}"

                lista_mes = next(
                    (lst for lst in lists if termo_busca in normalizar(lst.get("name", ""))),
                    None
                )
                lista_fallback = next(
                    (lst for lst in lists if "EM EXECUCAO" in normalizar(lst.get("name", ""))),
                    None
                )
                lista_alvo = lista_mes or lista_fallback

                if lista_alvo:
                    id_lista_execucao = lista_alvo.get("id")
                    nome_lista_alvo = lista_alvo.get("name", "")
            else:
                lista_alvo = next(
                    (lst for lst in lists if lst.get("id") == id_lista_execucao),
                    None
                )
                if lista_alvo:
                    nome_lista_alvo = lista_alvo.get("name", "")

            if not id_lista_execucao:
                return False, "Nenhuma lista do mês vigente ou de execução foi encontrada no Trello."

            cards_execucao = [
                c for c in cards
                if c.get("idList") == id_lista_execucao and not c.get("closed", False)
            ]
            nome_alvo = f"Lista: {nome_lista_alvo}"

        # IMPORTANTE:
        # Não usamos buscar_obras() aqui, pois essa função geral retorna [] quando há
        # falha de conexão. Durante uma sincronização isso poderia parecer "banco vazio"
        # e fazer o sistema tentar reinserir todas as obras.
        try:
            resposta_obras = _executar_supabase_com_retry(
                lambda sb: sb.table("obras").select("id,nome,unidade").execute(),
                tentativas=3
            )
            obras_atuais = resposta_obras.data or []
        except Exception as e:
            try:
                st.session_state["supabase_ultimo_erro_sync"] = (
                    f"{type(e).__name__}: {str(e)[:220]}"
                )
            except Exception:
                pass

            return False, (
                "Não foi possível conectar ao banco de dados agora. "
                "Nenhuma obra foi alterada. Aguarde alguns segundos e tente sincronizar novamente."
            )

        nomes_cadastrados = {
            normalizar(o.get("nome", ""))
            for o in obras_atuais
            if o.get("nome")
        }

        novas_inseridas = 0
        ja_existentes = 0
        falhas = []

        for card in cards_execucao:
            nome_card = str(card.get("name", "") or "").strip()
            if not nome_card:
                continue

            nome_norm = normalizar(nome_card)
            unidade_card = identificar_unidade(nome_card)

            if nome_norm in nomes_cadastrados:
                ja_existentes += 1
                continue

            ok, situacao = _inserir_obra_resiliente(
                nome_obra=nome_card,
                unidade=unidade_card,
                tentativas=3
            )

            if not ok:
                falhas.append(nome_card)
                continue

            nomes_cadastrados.add(nome_norm)

            if situacao == "inserida":
                novas_inseridas += 1
            else:
                # Pode ter sido criada em uma tentativa anterior cuja resposta se perdeu,
                # ou já existir no banco apesar do snapshot inicial.
                ja_existentes += 1

        limpar_cache_operacional()

        if falhas:
            return False, (
                f"Sincronização de {nome_alvo} concluída parcialmente: "
                f"{novas_inseridas} nova(s) obra(s) adicionada(s), "
                f"{ja_existentes} já existente(s) e "
                f"{len(falhas)} item(ns) não puderam ser confirmados no banco. "
                "Tente sincronizar novamente em alguns segundos."
            )

        return True, (
            f"Sincronização de {nome_alvo} concluída: "
            f"{novas_inseridas} nova(s) obra(s) adicionada(s) e "
            f"{ja_existentes} já existente(s)."
        )

    except Exception as e:
        # Última barreira: nenhum erro da sincronização deve abrir o traceback vermelho.
        try:
            st.session_state["supabase_ultimo_erro_sync"] = (
                f"{type(e).__name__}: {str(e)[:220]}"
            )
        except Exception:
            pass

        return False, (
            "Não foi possível concluir a sincronização agora. "
            "O sistema continua funcionando com os dados já cadastrados. "
            "Aguarde alguns segundos e tente novamente."
        )


# --- BUSCA DE DADOS COM CACHE ---
@st.cache_data(ttl=30)
def buscar_obras():
    try: return supabase.table("obras").select("*").execute().data
    except Exception: return []

@st.cache_data(ttl=30)
def buscar_colaboradores():
    try: 
        res = supabase.table("colaboradores").select("*").execute().data
        return res if res else []
    except Exception: 
        return []

obras = buscar_obras()
colaboradores = buscar_colaboradores()

dict_colaboradores = {c['id']: c for c in colaboradores} if colaboradores else {}
dict_obras = {o['id']: o for o in obras} if obras else {}

ENGENHEIROS = ["EDUARDO", "GABRIEL", "GUSTAVO", "JOEL", "NETO", "PAULO", "SOARES", "VICTOR"]

UNIDADES_APROAR = [
    "BARRA DO CEARÁ",
    "MARACANAÚ",
    "COLISEU",
    "HORIZONTE",
    "ESCRITÓRIO",
    "CENTRO",
    "MUSEU",
    "FIEC",
    "UNIFOR",
    "SEBRAE",
]

# --- FUNÇÃO AUXILIAR PARA RENDERIZAR A ABA DE DISPONIBILIDADE ---
def render_aba_disponibilidade(key_suffix=""):
    st.markdown("### 👥 DISPONIBILIDADE DE EQUIPE POR FUNÇÃO")
    st.write("Consulte quem já está convocado e quem está disponível para a data selecionada.")
    
    data_disp = st.date_input("Data de referência:", value=datetime.date.today() + datetime.timedelta(days=1), format="DD/MM/YYYY", key=f"data_disp_{key_suffix}")
    
    try:
        convs_disp = supabase.table("convocacoes").select("*").eq("data", data_disp.isoformat()).execute().data
    except:
        convs_disp = []
        
    ids_ocupados = {c['colaborador_id'] for c in convs_disp}
    funcoes = sorted(list(set([c['funcao'] for c in colaboradores])))
    
    if not funcoes:
        st.info("Nenhum colaborador cadastrado.")
        return

    st.markdown("---")
    for func in funcoes:
        with st.container(border=True):
            st.markdown(f"#### 🔹 {func.upper()}")
            colabs_func = [c for c in colaboradores if c['funcao'] == func]
            ocupados_func = [c for c in colabs_func if c['id'] in ids_ocupados]
            disponiveis_func = [c for c in colabs_func if c['id'] not in ids_ocupados]
            
            c1, c2 = st.columns(2)
            with c1:
                st.markdown(f"**🔴 CONVOCADOS ({len(ocupados_func)})**")
                if ocupados_func:
                    for oc in ocupados_func:
                        conv_info = next((item for item in convs_disp if item['colaborador_id'] == oc['id']), None)
                        obra_nome = "Obra"
                        eng_resp = ""
                        if conv_info:
                            ob_inf = dict_obras.get(conv_info['obra_id'], {})
                            obra_nome = f"{ob_inf.get('unidade','')} - {ob_inf.get('nome','')}"
                            eng_resp = f" (Eng: {conv_info.get('engenheiro', 'N/A')})"
                        st.markdown(f"• {oc['nome']} <br><small style='color:#94A3B8;'>({obra_nome}{eng_resp})</small>", unsafe_allow_html=True)
                else:
                    st.caption("Nenhum.")
                    
            with c2:
                st.markdown(f"**🟢 DISPONÍVEIS ({len(disponiveis_func)})**")
                if disponiveis_func:
                    for disp in disponiveis_func:
                        st.markdown(f"• {disp['nome']}")
                else:
                    st.caption("Nenhum disponível.")




# --- MELHORIAS OPERACIONAIS / RAMON ---
INDISP_UNIDADE = "__APROAR_INDISPONIBILIDADE__"
INDISP_PREFIX = "APROAR_INDISP|"


def eh_registro_indisponibilidade(obra):
    return bool(obra) and (
        str(obra.get("unidade") or "") == INDISP_UNIDADE
        or str(obra.get("nome") or "").startswith(INDISP_PREFIX)
    )


def decodificar_indisponibilidade(obra):
    if not eh_registro_indisponibilidade(obra):
        return None
    try:
        bruto = str(obra.get("nome") or "")[len(INDISP_PREFIX):]
        dados = json.loads(bruto)
        if not isinstance(dados, dict):
            return None
        dados["id"] = obra.get("id")
        return dados
    except Exception:
        return None


def listar_indisponibilidades():
    saida = []
    for item in obras_todas:
        dados = decodificar_indisponibilidade(item)
        if dados:
            saida.append(dados)
    return saida


def obter_indisponibilidade_colaborador(colaborador_id, data_ref):
    if isinstance(data_ref, datetime.datetime):
        data_ref = data_ref.date()
    for item in listar_indisponibilidades():
        if str(item.get("colaborador_id")) != str(colaborador_id):
            continue
        try:
            ini = datetime.date.fromisoformat(str(item.get("inicio")))
            fim = datetime.date.fromisoformat(str(item.get("fim")))
        except Exception:
            continue
        if ini <= data_ref <= fim:
            return item
    return None


def salvar_indisponibilidade(colaborador_id, motivo, inicio, fim, observacao=""):
    if fim < inicio:
        return False, "A data final não pode ser anterior à data inicial."
    dados = {
        "colaborador_id": str(colaborador_id),
        "motivo": str(motivo),
        "inicio": inicio.isoformat(),
        "fim": fim.isoformat(),
        "observacao": str(observacao or "").strip(),
        "criado_em": agora_aproar().isoformat(),
    }
    try:
        supabase.table("obras").insert({
            "unidade": INDISP_UNIDADE,
            "nome": INDISP_PREFIX + json.dumps(dados, ensure_ascii=False, separators=(",", ":")),
        }).execute()
        limpar_cache_operacional()
        return True, "Indisponibilidade registrada."
    except Exception as e:
        return False, f"Não foi possível registrar a indisponibilidade: {e}"


def excluir_indisponibilidade(registro_id):
    try:
        supabase.table("obras").delete().eq("id", registro_id).execute()
        limpar_cache_operacional()
        return True
    except Exception:
        return False


obras_todas = buscar_obras() or []
obras = [o for o in obras_todas if not eh_registro_indisponibilidade(o)]
colaboradores = buscar_colaboradores() or []

dict_colaboradores = {c['id']: c for c in colaboradores} if colaboradores else {}
dict_obras = {o['id']: o for o in obras} if obras else {}

ENGENHEIROS = ["EDUARDO", "GABRIEL", "GUSTAVO", "JOEL", "NETO", "PAULO", "SOARES", "VICTOR"]

# --- FUNÇÃO AUXILIAR PARA RENDERIZAR A ABA DE DISPONIBILIDADE ---
def render_aba_disponibilidade(key_suffix=""):
    st.markdown("### 👥 Disponibilidade da equipe")
    st.caption("Convocados, indisponíveis (férias/atestados/afastamentos) e pessoas livres para a data escolhida.")

    data_disp = st.date_input(
        "Data de referência:",
        value=proximo_dia_util(agora_aproar().date()),
        format="DD/MM/YYYY",
        key=f"data_disp_{key_suffix}",
    )
    try:
        convs_disp = supabase.table("convocacoes").select("*").eq("data", data_disp.isoformat()).execute().data or []
    except Exception:
        convs_disp = []

    ids_convocados = {str(c.get("colaborador_id")) for c in convs_disp}
    indisponiveis_map = {}
    for c in colaboradores:
        ind = obter_indisponibilidade_colaborador(c.get("id"), data_disp)
        if ind:
            indisponiveis_map[str(c.get("id"))] = ind

    funcoes = sorted({str(c.get("funcao") or "INDEFINIDA") for c in colaboradores})
    if not funcoes:
        st.info("Nenhum colaborador cadastrado.")
        return

    for func in funcoes:
        colabs_func = [c for c in colaboradores if str(c.get("funcao") or "INDEFINIDA") == func]
        convocados = [c for c in colabs_func if str(c.get("id")) in ids_convocados]
        indisponiveis = [c for c in colabs_func if str(c.get("id")) not in ids_convocados and str(c.get("id")) in indisponiveis_map]
        disponiveis = [c for c in colabs_func if str(c.get("id")) not in ids_convocados and str(c.get("id")) not in indisponiveis_map]

        with st.container(border=True):
            st.markdown(f"#### {func.upper()}")
            c1, c2, c3 = st.columns(3)
            with c1:
                st.markdown(f"**🔵 CONVOCADOS ({len(convocados)})**")
                for oc in convocados:
                    conv = next((x for x in convs_disp if str(x.get("colaborador_id")) == str(oc.get("id"))), None)
                    ob = dict_obras.get((conv or {}).get("obra_id"), {})
                    st.markdown(f"• **{oc.get('nome','-')}**")
                    st.caption(f"{ob.get('unidade','-')} • Eng. {(conv or {}).get('engenheiro','-')}")
                if not convocados:
                    st.caption("Nenhum.")
            with c2:
                st.markdown(f"**🔴 INDISPONÍVEIS ({len(indisponiveis)})**")
                for oc in indisponiveis:
                    ind = indisponiveis_map.get(str(oc.get("id")), {})
                    st.markdown(f"• **{oc.get('nome','-')}**")
                    st.caption(f"{ind.get('motivo','Indisponível')} • {ind.get('inicio','')} a {ind.get('fim','')}")
                if not indisponiveis:
                    st.caption("Nenhum.")
            with c3:
                st.markdown(f"**🟢 DISPONÍVEIS ({len(disponiveis)})**")
                for oc in disponiveis:
                    st.markdown(f"• {oc.get('nome','-')}")
                if not disponiveis:
                    st.caption("Nenhum.")


# --- COMPONENTES COMPARTILHADOS: APONTAMENTO, DASHBOARD, RELATÓRIO E AUDITORIA ---
OPCOES_STATUS_PRESENCA = [
    "Presente (Integral)", "Presente (Só Manhã)", "Presente (Só Tarde)",
    "Saída Antecipada", "Falta", "Atestado"
]


def _buscar_convocacoes_intervalo(data_inicio, data_fim, engenheiro=None):
    try:
        q = supabase.table("convocacoes").select("*").gte("data", data_inicio.isoformat()).lte("data", data_fim.isoformat())
        if engenheiro:
            q = q.eq("engenheiro", engenheiro)
        return q.execute().data or []
    except Exception:
        return []


def _periodo_por_tipo(tipo, data_base):
    if tipo == "Diário":
        return data_base, data_base
    if tipo == "Semanal":
        inicio = data_base - datetime.timedelta(days=data_base.weekday())
        return inicio, inicio + datetime.timedelta(days=6)
    if tipo == "Mensal":
        inicio = data_base.replace(day=1)
        if inicio.month == 12:
            prox = datetime.date(inicio.year + 1, 1, 1)
        else:
            prox = datetime.date(inicio.year, inicio.month + 1, 1)
        return inicio, prox - datetime.timedelta(days=1)
    return data_base, data_base


def _processar_registro_operacional(registro):
    obra = dict_obras.get(registro.get("obra_id"), {"unidade": "GERAL", "nome": "Desconhecida"})
    colab = dict_colaboradores.get(registro.get("colaborador_id"), {"nome": "Desconhecido", "funcao": "-"})
    status = normalizar_status_operacional(registro.get("status"))
    diaria = calcular_diaria_proporcional(status, obter_valor_diaria_colaborador(colab))
    extra_bruta = float(registro.get("valor_extra") or 0.0)
    extra = extra_bruta if status_eh_presenca(status) else 0.0
    meta = obter_metadata_operacional(registro.get("observacao") or "")
    _, obs_livre = decompor_observacao_operacional(registro.get("observacao") or "")
    return {
        "id": registro.get("id"),
        "Data": str(registro.get("data") or ""),
        "Engenheiro": str(registro.get("engenheiro") or "N/A"),
        "Unidade": str(obra.get("unidade") or "GERAL"),
        "Serviço(s)": descricao_servicos_convocacao(registro, obra),
        "Colaborador": str(colab.get("nome") or "Desconhecido"),
        "Função": str(colab.get("funcao") or "-"),
        "Status": status,
        "Diária (R$)": float(diaria),
        "Extra (R$)": extra,
        "Custo (R$)": float(diaria + extra),
        "Observação": obs_livre,
        "Convocação após 16h": bool(meta.get("convocacao_atrasada")),
        "Apontamento atrasado": bool(meta.get("apontamento_atrasado")),
        "Apontado em": str(meta.get("apontado_em") or ""),
    }


def _gerar_excel_dataframe(df, titulo="APROAR - RELATÓRIO"):
    buffer = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relatório"
    total_cols = max(1, len(df.columns))
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    c = ws.cell(1, 1, titulo)
    c.font = Font(name="Arial", size=13, bold=True, color="FFFFFF")
    c.fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    c.alignment = Alignment(horizontal="center")
    for ci, col in enumerate(df.columns, 1):
        cell = ws.cell(3, ci, str(col))
        cell.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1D4ED8", end_color="1D4ED8", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    for ri, (_, row) in enumerate(df.iterrows(), 4):
        for ci, col in enumerate(df.columns, 1):
            val = row[col]
            if pd.isna(val):
                val = ""
            cell = ws.cell(ri, ci, val)
            cell.font = Font(name="Arial", size=9)
            if "(R$)" in str(col) and isinstance(val, (int, float)):
                cell.number_format = 'R$ #,##0.00'
    for ci, col in enumerate(df.columns, 1):
        amostra = [len(str(col))] + [len(str(v)) for v in df[col].head(100).tolist()]
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = min(max(amostra) + 3, 42)
    ws.freeze_panes = "A4"
    wb.save(buffer)
    return buffer.getvalue()


def gerar_excel_dashboard_consolidado(df, tipo, inicio, fim):
    """Excel do Dashboard com resumo, custo por Unidade/Engenheiro e detalhamento."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumo"

    total = len(df)
    presentes = int(df["Status"].apply(status_eh_presenca).sum()) if not df.empty else 0
    faltas = int((df["Status"] == "Falta").sum()) if not df.empty else 0
    atestados = int((df["Status"] == "Atestado").sum()) if not df.empty else 0
    custo = float(df["Custo (R$)"].sum()) if not df.empty else 0.0
    extras = float(df["Extra (R$)"].sum()) if not df.empty else 0.0
    taxa = (presentes / total * 100) if total else 0.0

    ws["A1"] = f"APROAR - DASHBOARD {str(tipo).upper()}"
    ws["A1"].font = Font(name="Arial", size=13, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    ws.merge_cells("A1:B1")
    ws["A2"] = f"Período: {inicio.strftime('%d/%m/%Y')} a {fim.strftime('%d/%m/%Y')}"
    resumo = [
        ("Convocados / registros", total),
        ("Presentes", presentes),
        ("Faltas", faltas),
        ("Atestados", atestados),
        ("Presença (%)", round(taxa, 1)),
        ("Extras (R$)", extras),
        ("Custo total (R$)", custo),
    ]
    for i, (rotulo, valor) in enumerate(resumo, 4):
        ws.cell(i, 1, rotulo).font = Font(name="Arial", size=10, bold=True)
        ws.cell(i, 2, valor)
        if "(R$)" in rotulo:
            ws.cell(i, 2).number_format = 'R$ #,##0.00'
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18

    def add_df_sheet(nome, dados):
        sh = wb.create_sheet(nome)
        if dados is None or dados.empty:
            sh["A1"] = "Sem dados"
            return
        for ci, col in enumerate(dados.columns, 1):
            cell = sh.cell(1, ci, str(col))
            cell.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1D4ED8", end_color="1D4ED8", fill_type="solid")
        for ri, (_, row) in enumerate(dados.iterrows(), 2):
            for ci, col in enumerate(dados.columns, 1):
                val = row[col]
                if pd.isna(val):
                    val = ""
                cell = sh.cell(ri, ci, val)
                if "(R$)" in str(col) and isinstance(val, (int, float)):
                    cell.number_format = 'R$ #,##0.00'
        for ci, col in enumerate(dados.columns, 1):
            vals = [len(str(col))] + [len(str(v)) for v in dados[col].head(200).tolist()]
            sh.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = min(max(vals) + 3, 42)
        sh.freeze_panes = "A2"

    por_unidade = df.groupby("Unidade", dropna=False)["Custo (R$)"].sum().reset_index().sort_values("Custo (R$)", ascending=False) if not df.empty else pd.DataFrame()
    por_eng = df.groupby("Engenheiro", dropna=False)["Custo (R$)"].sum().reset_index().sort_values("Custo (R$)", ascending=False) if not df.empty else pd.DataFrame()
    add_df_sheet("Custo por Unidade", por_unidade)
    add_df_sheet("Custo por Engenheiro", por_eng)
    add_df_sheet("Detalhamento", df)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def gerar_excel_indicador_prazos(df_resumo, df_eventos, inicio, fim):
    """Relatório de cumprimento com resumo e as datas de cada evento auditável."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    def add(nome, dados, titulo):
        ws = wb.create_sheet(nome)
        total_cols = max(1, len(dados.columns))
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        ws.cell(1, 1, titulo).font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
        ws.cell(1, 1).fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        ws.cell(2, 1, f"Período: {inicio.strftime('%d/%m/%Y')} a {fim.strftime('%d/%m/%Y')}")
        if dados.empty:
            ws.cell(4, 1, "Sem dados")
            return
        for ci, col in enumerate(dados.columns, 1):
            c = ws.cell(4, ci, str(col))
            c.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
            c.fill = PatternFill(start_color="1D4ED8", end_color="1D4ED8", fill_type="solid")
        for ri, (_, row) in enumerate(dados.iterrows(), 5):
            for ci, col in enumerate(dados.columns, 1):
                val = row[col]
                if pd.isna(val):
                    val = ""
                ws.cell(ri, ci, val)
        for ci, col in enumerate(dados.columns, 1):
            vals = [len(str(col))] + [len(str(v)) for v in dados[col].head(200).tolist()]
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = min(max(vals) + 3, 42)
        ws.freeze_panes = "A5"

    add("Resumo", df_resumo, "APROAR - CUMPRIMENTO DE PRAZOS")
    add("Ocorrências", df_eventos, "APROAR - DATAS E OCORRÊNCIAS DE PRAZO")
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def render_dashboard_consulta(key_prefix="dash", engenheiro_fixo=None):
    st.markdown("## 🎛️ Dashboard")
    st.caption("Extra é valor adicional de quem esteve presente: entra no custo e não é contado como um status separado.")
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        tipo = st.selectbox("Período", ["Diário", "Semanal", "Mensal"], key=f"{key_prefix}_tipo")
    with c2:
        base = st.date_input("Data de referência", value=agora_aproar().date(), format="DD/MM/YYYY", key=f"{key_prefix}_base")
    inicio, fim = _periodo_por_tipo(tipo, base)
    with c3:
        unidades = sorted({o.get("unidade") for o in obras if o.get("unidade")})
        unidade = st.selectbox("Unidade", ["TODAS"] + unidades, key=f"{key_prefix}_unidade")
    with c4:
        if engenheiro_fixo:
            engenheiro = engenheiro_fixo
            st.text_input("Engenheiro", value=engenheiro_fixo, disabled=True, key=f"{key_prefix}_engfix")
        else:
            engenheiro = st.selectbox("Engenheiro", ["TODOS"] + ENGENHEIROS, key=f"{key_prefix}_eng")

    registros = _buscar_convocacoes_intervalo(inicio, fim, None if engenheiro == "TODOS" else engenheiro)
    processados = []
    for r in registros:
        item = _processar_registro_operacional(r)
        if unidade != "TODAS" and item["Unidade"] != unidade:
            continue
        processados.append(item)

    total = len(processados)
    presentes = sum(1 for x in processados if status_eh_presenca(x["Status"]))
    faltas = sum(1 for x in processados if x["Status"] == "Falta")
    atestados = sum(1 for x in processados if x["Status"] == "Atestado")
    custo = sum(float(x["Custo (R$)"]) for x in processados)
    total_extra = sum(float(x["Extra (R$)"]) for x in processados)
    taxa_presenca = (presentes / total * 100) if total else 0.0

    m1, m2, m3, m4, m5 = st.columns(5)
    m1.metric("CONVOCADOS / REGISTROS", total)
    m2.metric("PRESENTES", presentes)
    m3.metric("FALTAS", faltas)
    m4.metric("ATESTADOS", atestados)
    m5.metric("CUSTO TOTAL", formatar_reais(custo))
    st.caption(
        f"Período: {inicio.strftime('%d/%m/%Y')} a {fim.strftime('%d/%m/%Y')} • "
        f"Presença: {taxa_presenca:.1f}% • Extras incluídas no custo: {formatar_reais(total_extra)}"
    )

    if not processados:
        st.info("Nenhum registro encontrado para os filtros selecionados.")
        return

    df = pd.DataFrame(processados)

    st.markdown("### 💰 Custos consolidados")
    g1, g2 = st.columns(2)
    with g1:
        st.markdown("#### Por Unidade")
        por_unidade = (
            df.groupby("Unidade", dropna=False)["Custo (R$)"]
            .sum().reset_index().sort_values("Custo (R$)", ascending=False)
        )
        st.dataframe(por_unidade, use_container_width=True, hide_index=True)
        if not por_unidade.empty:
            st.bar_chart(por_unidade.set_index("Unidade")["Custo (R$)"], use_container_width=True)
    with g2:
        st.markdown("#### Por Engenheiro")
        por_eng = (
            df.groupby("Engenheiro", dropna=False)["Custo (R$)"]
            .sum().reset_index().sort_values("Custo (R$)", ascending=False)
        )
        st.dataframe(por_eng, use_container_width=True, hide_index=True)
        if not por_eng.empty:
            st.bar_chart(por_eng.set_index("Engenheiro")["Custo (R$)"], use_container_width=True)

    st.markdown("### 📋 Detalhamento")
    cols = ["Data", "Engenheiro", "Unidade", "Serviço(s)", "Colaborador", "Status", "Diária (R$)", "Extra (R$)", "Custo (R$)"]
    st.dataframe(df[cols], use_container_width=True, hide_index=True)
    st.download_button(
        "📥 BAIXAR DASHBOARD EM EXCEL",
        data=gerar_excel_dashboard_consolidado(df, tipo, inicio, fim),
        file_name=f"dashboard_{tipo.lower()}_{inicio.isoformat()}_a_{fim.isoformat()}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        key=f"{key_prefix}_download",
    )


def render_relatorio_visualizador(key_prefix="rel_view", engenheiro_fixo=None):
    st.markdown("## 📊 Relatórios")
    c1, c2, c3 = st.columns(3)
    with c1:
        inicio = st.date_input("Início", value=agora_aproar().date().replace(day=1), format="DD/MM/YYYY", key=f"{key_prefix}_ini")
    with c2:
        fim = st.date_input("Fim", value=agora_aproar().date(), format="DD/MM/YYYY", key=f"{key_prefix}_fim")
    with c3:
        unidades = sorted({o.get("unidade") for o in obras if o.get("unidade")})
        unidade = st.selectbox("Unidade", ["TODAS"] + unidades, key=f"{key_prefix}_unid")
    if inicio > fim:
        st.error("A data inicial não pode ser maior que a final.")
        return
    registros = _buscar_convocacoes_intervalo(inicio, fim, engenheiro_fixo)
    linhas = []
    for r in registros:
        p = _processar_registro_operacional(r)
        if unidade != "TODAS" and p["Unidade"] != unidade:
            continue
        linhas.append(p)
    if not linhas:
        st.info("Sem registros para o período.")
        return
    df = pd.DataFrame(linhas)
    total = float(df["Custo (R$)"].sum())
    e1, e2, e3 = st.columns(3)
    e1.metric("CUSTO TOTAL", formatar_reais(total))
    e2.metric("PESSOAS/DIA", len(df))
    e3.metric("DIAS COM REGISTRO", df["Data"].nunique())
    st.dataframe(df[["Data", "Unidade", "Serviço(s)", "Colaborador", "Status", "Diária (R$)", "Extra (R$)", "Custo (R$)"]], use_container_width=True, hide_index=True)
    st.download_button(
        "📥 BAIXAR RELATÓRIO EXCEL",
        data=_gerar_excel_dataframe(df, "APROAR - RELATÓRIO DE APONTAMENTOS"),
        file_name=f"relatorio_apontamentos_{inicio.isoformat()}_a_{fim.isoformat()}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        key=f"{key_prefix}_download",
    )


def render_indicadores_cumprimento(key_prefix="ind", engenheiro_fixo=None, mostrar_absenteismo=True):
    st.markdown("## 📈 Indicadores")
    c1, c2, c3 = st.columns(3)
    with c1:
        inicio = st.date_input("Início", value=agora_aproar().date() - datetime.timedelta(days=30), format="DD/MM/YYYY", key=f"{key_prefix}_ini")
    with c2:
        fim = st.date_input("Fim", value=agora_aproar().date(), format="DD/MM/YYYY", key=f"{key_prefix}_fim")
    with c3:
        unidades = sorted({o.get("unidade") for o in obras if o.get("unidade")})
        unidade_filtro = st.selectbox("Unidade", ["TODAS"] + unidades, key=f"{key_prefix}_unidade")
    if inicio > fim:
        st.error("Período inválido.")
        return

    registros_brutos = _buscar_convocacoes_intervalo(inicio, fim, engenheiro_fixo)
    registros = []
    eventos_prazo = []

    for r in registros_brutos:
        obra = dict_obras.get(r.get("obra_id"), {"unidade": "GERAL"})
        if unidade_filtro != "TODAS" and obra.get("unidade") != unidade_filtro:
            continue
        colab = dict_colaboradores.get(r.get("colaborador_id"), {"nome": "Desconhecido", "valor_diaria": VALOR_DIARIA_PROFISSIONAL})
        status = normalizar_status_operacional(r.get("status"))
        meta = obter_metadata_operacional(r.get("observacao") or "")
        try:
            data_dt = pd.to_datetime(r.get("data"), errors="coerce")
        except Exception:
            data_dt = pd.NaT
        eng = str(r.get("engenheiro") or "N/A")
        nome_colab = str(colab.get("nome") or "Desconhecido")
        registros.append({
            "raw": r,
            "engenheiro": eng,
            "unidade": str(obra.get("unidade") or "GERAL"),
            "colaborador": nome_colab,
            "status": status,
            "valor_diaria": obter_valor_diaria_colaborador(colab),
            "data": data_dt,
            "meta": meta,
        })

        if meta.get("convocado_em"):
            eventos_prazo.append({
                "Engenheiro": eng,
                "Data do serviço": str(r.get("data") or ""),
                "Colaborador": nome_colab,
                "Tipo": "Convocação",
                "Registrado em": str(meta.get("convocado_em") or "").replace("T", " ")[:19],
                "Atrasado": "SIM" if meta.get("convocacao_atrasada") else "NÃO",
            })
        if meta.get("apontado_em"):
            eventos_prazo.append({
                "Engenheiro": eng,
                "Data do serviço": str(r.get("data") or ""),
                "Colaborador": nome_colab,
                "Tipo": "Apontamento",
                "Registrado em": str(meta.get("apontado_em") or "").replace("T", " ")[:19],
                "Atrasado": "SIM" if meta.get("apontamento_atrasado") else "NÃO",
            })

    if not registros:
        st.info("Sem registros no período.")
        return

    por_eng = {}
    for item in registros:
        eng = item["engenheiro"]
        meta = item["meta"]
        d = por_eng.setdefault(eng, {
            "Engenheiro": eng,
            "Convocações auditáveis": 0,
            "Convocações atrasadas": 0,
            "Apontamentos auditáveis": 0,
            "Apontamentos atrasados": 0,
        })
        if meta.get("convocado_em"):
            d["Convocações auditáveis"] += 1
            d["Convocações atrasadas"] += int(bool(meta.get("convocacao_atrasada")))
        if meta.get("apontado_em"):
            d["Apontamentos auditáveis"] += 1
            d["Apontamentos atrasados"] += int(bool(meta.get("apontamento_atrasado")))

    linhas = []
    for d in por_eng.values():
        auditaveis = d["Convocações auditáveis"] + d["Apontamentos auditáveis"]
        atrasos = d["Convocações atrasadas"] + d["Apontamentos atrasados"]
        d["No prazo (%)"] = round(((auditaveis - atrasos) / auditaveis * 100), 1) if auditaveis else None
        linhas.append(d)
    df_prazos = pd.DataFrame(linhas).sort_values("Engenheiro")

    st.markdown("### ⏱️ Cumprimento por engenheiro")
    st.caption("Convocação atrasada = feita após 16h para o próximo dia útil. Apontamento atrasado = salvo em dia posterior ao serviço. Os dois atrasos são medidos separadamente.")
    st.dataframe(df_prazos, use_container_width=True, hide_index=True)

    # Detalhamento clicável/selecionável das datas que geraram atraso.
    df_eventos = pd.DataFrame(eventos_prazo)
    if not df_eventos.empty:
        st.markdown("#### 🔎 Ver datas e ocorrências")
        op_eng = sorted(df_eventos["Engenheiro"].dropna().astype(str).unique().tolist())
        eng_det = engenheiro_fixo or st.selectbox("Engenheiro para detalhar", op_eng, key=f"{key_prefix}_eng_detalhe")
        somente_atrasos = st.checkbox("Mostrar somente atrasos", value=True, key=f"{key_prefix}_somente_atrasos")
        det = df_eventos[df_eventos["Engenheiro"] == eng_det].copy()
        if somente_atrasos:
            det = det[det["Atrasado"] == "SIM"]
        if det.empty:
            st.success("Nenhuma ocorrência atrasada para este engenheiro no período.")
        else:
            st.dataframe(det.sort_values(["Data do serviço", "Tipo"], ascending=[False, True]), use_container_width=True, hide_index=True)

    st.download_button(
        "📥 BAIXAR INDICADOR DE PRAZOS",
        data=gerar_excel_indicador_prazos(df_prazos, df_eventos, inicio, fim),
        file_name=f"indicador_prazos_{inicio.isoformat()}_a_{fim.isoformat()}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key=f"{key_prefix}_download",
    )

    if not mostrar_absenteismo:
        return

    df = pd.DataFrame([{k: v for k, v in x.items() if k not in ["raw", "meta"]} for x in registros])
    total_conv = len(df)
    total_faltas = int((df["status"] == "Falta").sum())
    total_atestados = int((df["status"] == "Atestado").sum())
    total_ausencias = total_faltas + total_atestados
    taxa_absenteismo = (total_ausencias / total_conv * 100) if total_conv else 0.0
    mask_ausencia = df["status"].isin(["Falta", "Atestado"])
    impacto_financeiro = float(df.loc[mask_ausencia, "valor_diaria"].sum())

    st.markdown("---")
    st.markdown("### 👥 Absenteísmo")
    m1, m2, m3, m4, m5 = st.columns(5)
    m1.metric("CONVOCAÇÕES", total_conv)
    m2.metric("FALTAS", total_faltas)
    m3.metric("ATESTADOS", total_atestados)
    m4.metric("ABSENTEÍSMO", f"{taxa_absenteismo:.1f}%")
    m5.metric("IMPACTO EST.", formatar_reais(impacto_financeiro))
    st.caption("Impacto estimado = soma das diárias-base associadas às faltas e atestados.")

    r1, r2 = st.columns(2)
    with r1:
        st.markdown("#### Colaboradores com mais faltas")
        df_faltas = df[df["status"] == "Falta"]
        if df_faltas.empty:
            st.info("Nenhuma falta registrada no período.")
        else:
            ranking = (
                df_faltas.groupby("colaborador").size().reset_index(name="Faltas")
                .sort_values(["Faltas", "colaborador"], ascending=[False, True]).reset_index(drop=True)
            )
            ranking.insert(0, "Posição", range(1, len(ranking) + 1))
            st.dataframe(ranking, use_container_width=True, hide_index=True)

    with r2:
        st.markdown("#### Ausências por dia da semana")
        dias_ordem = ["Segunda", "Terça", "Quarta", "Quinta", "Sexta", "Sábado", "Domingo"]
        mapa_dias = {0: "Segunda", 1: "Terça", 2: "Quarta", 3: "Quinta", 4: "Sexta", 5: "Sábado", 6: "Domingo"}
        df_aus = df[mask_ausencia].copy()
        if df_aus.empty or df_aus["data"].isna().all():
            st.info("Sem ausências com data válida.")
        else:
            df_aus = df_aus.dropna(subset=["data"])
            df_aus["Dia"] = df_aus["data"].dt.weekday.map(mapa_dias)
            resumo_semana = df_aus.groupby(["Dia", "status"]).size().unstack(fill_value=0).reindex(dias_ordem, fill_value=0)
            for coluna in ["Falta", "Atestado"]:
                if coluna not in resumo_semana.columns:
                    resumo_semana[coluna] = 0
            st.bar_chart(resumo_semana[["Falta", "Atestado"]], use_container_width=True)

    st.markdown("#### Detalhamento por Unidade")
    resumo_unidades = []
    for und in sorted(df["unidade"].unique()):
        df_u = df[df["unidade"] == und]
        t_u = len(df_u)
        f_u = int((df_u["status"] == "Falta").sum())
        a_u = int((df_u["status"] == "Atestado").sum())
        aus_u = f_u + a_u
        taxa_u = (aus_u / t_u * 100) if t_u else 0.0
        impacto_u = float(df_u.loc[df_u["status"].isin(["Falta", "Atestado"]), "valor_diaria"].sum())
        resumo_unidades.append({
            "Unidade": und,
            "Convocações": t_u,
            "Faltas": f_u,
            "Atestados": a_u,
            "Total Ausências": aus_u,
            "Taxa Absenteísmo (%)": round(taxa_u, 1),
            "Impacto Estimado (R$)": round(impacto_u, 2),
        })
    st.dataframe(pd.DataFrame(resumo_unidades).sort_values("Taxa Absenteísmo (%)", ascending=False), use_container_width=True, hide_index=True)


def incluir_colaborador_direto_apontamento(colaborador_id, engenheiro, data_servico, obra_id, turno="Integral"):
    ind = obter_indisponibilidade_colaborador(colaborador_id, data_servico)
    if ind:
        return False, f"Colaborador indisponível: {ind.get('motivo','Indisponível')} ({ind.get('inicio')} a {ind.get('fim')})."
    existente = buscar_convocacao_existente(colaborador_id, data_servico)
    if existente:
        reg = existente[0]
        if normalizar(reg.get("engenheiro")) != normalizar(engenheiro):
            registrar_conflito_convocacao(reg, engenheiro)
            return False, f"Esse colaborador já está com {reg.get('engenheiro','outro engenheiro')}. O conflito foi registrado para o Paulo."
        return False, "Esse colaborador já está no seu apontamento desta data."
    agora = agora_aproar()
    meta = {
        "convocado_em": agora.isoformat(),
        "convocado_por": str(engenheiro),
        "incluido_direto_apontamento": True,
        "convocacao_atrasada": bool(agora.hour >= 16 and data_servico == proximo_dia_util(agora.date())),
        "apontado_em": agora.isoformat(),
        "ultimo_apontamento_em": agora.isoformat(),
        "apontado_por": str(engenheiro),
        "apontamento_atrasado": bool(agora.date() > data_servico),
        "servicos_extras": [],
        "servicos_adicionais": [],
        "periodo_servico_principal": turno,
    }
    try:
        supabase.table("convocacoes").insert({
            "obra_id": obra_id,
            "colaborador_id": colaborador_id,
            "data": data_servico.isoformat(),
            "engenheiro": engenheiro,
            "status": "Presente (Integral)",
            "valor_extra": 0,
            "observacao": montar_observacao_operacional(turno, "", meta),
        }).execute()
        limpar_cache_operacional()
        return True, "Colaborador incluído no apontamento."
    except Exception:
        return False, "Não foi possível incluir o colaborador."


def render_apontamento_operacional(engenheiro_fixo=None, key_prefix="apont"):
    st.markdown("## ✅ Apontamento diário")
    st.caption("Presença e extra são independentes. O mesmo colaborador pode atuar em mais de um serviço da mesma Unidade sem duplicar convocação nem diária.")
    c1, c2, c3 = st.columns(3)
    with c1:
        if engenheiro_fixo:
            engenheiro = engenheiro_fixo
            st.text_input("Engenheiro", value=engenheiro, disabled=True, key=f"{key_prefix}_engfix")
        else:
            engenheiro = st.selectbox("Engenheiro", ENGENHEIROS, key=f"{key_prefix}_eng")
    with c2:
        data_apont = st.date_input("Data do serviço", value=agora_aproar().date(), format="DD/MM/YYYY", key=f"{key_prefix}_data")

    try:
        convs = supabase.table("convocacoes").select("*").eq("engenheiro", engenheiro).eq("data", data_apont.isoformat()).execute().data or []
    except Exception:
        convs = []
    for c in convs:
        c["dados_obra"] = dict_obras.get(c.get("obra_id"), {"unidade": "Desconhecida", "nome": NOME_OBRA_PLACEHOLDER})
    unidades_conv = sorted({(c.get("dados_obra") or {}).get("unidade", "Desconhecida") for c in convs})
    with c3:
        unidade_filtro = st.selectbox("Unidade", ["TODAS"] + unidades_conv, key=f"{key_prefix}_unidade")

    with st.expander("➕ Incluir colaborador que não estava na convocação", expanded=False):
        st.caption("A lista contém todos os colaboradores cadastrados e serve para correções ou inclusões excepcionais.")
        labels = {f"{c.get('nome')} ({c.get('funcao','-')})": c.get("id") for c in sorted(colaboradores, key=lambda x: normalizar(x.get('nome','')))}
        inc1, inc2 = st.columns(2)
        with inc1:
            nome_sel = st.selectbox("Colaborador", ["— Selecione —"] + list(labels.keys()), key=f"{key_prefix}_inc_colab")
        with inc2:
            unidades = sorted({o.get("unidade") for o in obras if o.get("unidade")})
            unid_inc = st.selectbox("Unidade", unidades, key=f"{key_prefix}_inc_unid") if unidades else None
        obras_inc = obras_reais_da_unidade(unid_inc) if unid_inc else []
        mapa_inc = {o.get("nome"): o.get("id") for o in obras_inc}
        obra_inc = st.selectbox("Obra / Serviço", ["— Selecione —"] + list(mapa_inc.keys()), key=f"{key_prefix}_inc_obra")
        if st.button("INCLUIR NO APONTAMENTO", type="primary", use_container_width=True, key=f"{key_prefix}_inc_btn"):
            if nome_sel == "— Selecione —" or obra_inc not in mapa_inc:
                st.warning("Selecione colaborador, Unidade e Obra/Serviço.")
            else:
                ok, msg = incluir_colaborador_direto_apontamento(labels[nome_sel], engenheiro, data_apont, mapa_inc[obra_inc])
                (st.success if ok else st.warning)(msg)
                if ok:
                    st.rerun()

    render = [c for c in convs if unidade_filtro == "TODAS" or (c.get("dados_obra") or {}).get("unidade") == unidade_filtro]
    if not render:
        st.info("Nenhuma equipe para os filtros selecionados.")
        return

    if st.button("✅ MARCAR EXIBIDOS COMO PRESENTE INTEGRAL", use_container_width=True, key=f"{key_prefix}_allpres"):
        for c in render:
            try:
                supabase.table("convocacoes").update({"status": "Presente (Integral)"}).eq("id", c.get("id")).execute()
            except Exception:
                pass
        st.rerun()

    periodos_servico = ["Integral", "Manhã", "Tarde", "Noite", "Outro"]

    for conv in render:
        c_id = conv.get("id")
        colab = dict_colaboradores.get(conv.get("colaborador_id"), {"nome": "Desconhecido", "funcao": "-"})
        unidade = (conv.get("dados_obra") or {}).get("unidade", "Desconhecida")
        obras_card = obras_reais_da_unidade(unidade)
        mapa_obras = {o.get("nome"): o.get("id") for o in obras_card}
        opcoes_obras = ["— Selecione a Obra/Serviço —"] + list(mapa_obras.keys())
        obra_atual = dict_obras.get(conv.get("obra_id"), {})
        nome_obra = obra_atual.get("nome", "")
        idx_obra = opcoes_obras.index(nome_obra) if nome_obra in mapa_obras else 0
        status_atual = normalizar_status_operacional(conv.get("status"))
        idx_st = OPCOES_STATUS_PRESENCA.index(status_atual) if status_atual in OPCOES_STATUS_PRESENCA else 0
        turno, obs_livre = decompor_observacao_operacional(conv.get("observacao") or "")
        meta_atual = obter_metadata_operacional(conv.get("observacao") or "")
        adicionais_atuais = [x for x in _normalizar_servicos_adicionais(meta_atual) if x.get("servico") in mapa_obras]
        atraso = bool(meta_atual.get("apontamento_atrasado"))
        periodo_principal_atual = str(meta_atual.get("periodo_servico_principal") or turno or "Integral")
        if periodo_principal_atual not in periodos_servico:
            periodo_principal_atual = "Outro"

        with st.container(border=True):
            st.markdown(f"### {colab.get('nome','-')}")
            st.caption(f"{colab.get('funcao','-')} • {unidade} • Convocado: {turno}")
            if atraso:
                st.warning("🟧 Apontamento realizado com atraso — salvo em data posterior ao serviço.")
            elif data_apont < agora_aproar().date() and not meta_atual.get("apontado_em"):
                st.warning("🟧 Este apontamento é retroativo. Ao salvar, o atraso será registrado.")

            with st.form(key=f"{key_prefix}_form_{c_id}"):
                f1, f2 = st.columns([1, 1.7])
                with f1:
                    status_sel = st.selectbox("Status", OPCOES_STATUS_PRESENCA, index=idx_st, key=f"{key_prefix}_st_{c_id}")
                with f2:
                    obra_sel = st.selectbox("Obra / Serviço principal", opcoes_obras, index=idx_obra, key=f"{key_prefix}_obra_{c_id}")

                idx_pp = periodos_servico.index(periodo_principal_atual)
                periodo_principal = st.selectbox("Período no serviço principal", periodos_servico, index=idx_pp, key=f"{key_prefix}_periodo_principal_{c_id}")

                st.markdown("**Outro serviço na mesma Unidade (opcional)**")
                opcoes_adic = ["— Nenhum —"] + list(mapa_obras.keys())
                adic_atual = adicionais_atuais[0]["servico"] if adicionais_atuais else "— Nenhum —"
                idx_adic = opcoes_adic.index(adic_atual) if adic_atual in opcoes_adic else 0
                s1, s2 = st.columns([1.7, 1])
                with s1:
                    segundo_servico = st.selectbox("2º serviço", opcoes_adic, index=idx_adic, key=f"{key_prefix}_seg_serv_{c_id}")
                with s2:
                    periodo_adic_atual = adicionais_atuais[0].get("periodo", "Tarde") if adicionais_atuais else "Tarde"
                    if periodo_adic_atual not in periodos_servico:
                        periodo_adic_atual = "Outro"
                    segundo_periodo = st.selectbox("Período do 2º serviço", periodos_servico, index=periodos_servico.index(periodo_adic_atual), key=f"{key_prefix}_seg_periodo_{c_id}")

                d1, d2 = st.columns([1, 2])
                with d1:
                    val_extra = st.number_input(
                        "Extra (R$)",
                        min_value=0.0,
                        value=(float(conv.get("valor_extra") or 0.0) if status_eh_presenca(status_sel) else 0.0),
                        step=10.0,
                        disabled=not status_eh_presenca(status_sel),
                        key=f"{key_prefix}_extra_{c_id}",
                        help="Extra é valor adicional. O colaborador continua contando como presente.",
                    )
                with d2:
                    obs_nova = st.text_input("Observação / justificativa", value=obs_livre, key=f"{key_prefix}_obs_{c_id}")
                salvar = st.form_submit_button("💾 SALVAR APONTAMENTO", type="primary", use_container_width=True)

            if salvar:
                if obra_sel not in mapa_obras:
                    st.warning("Selecione a Obra/Serviço principal antes de salvar.")
                elif segundo_servico == obra_sel:
                    st.warning("O 2º serviço deve ser diferente do serviço principal.")
                else:
                    adicionais = []
                    if segundo_servico in mapa_obras:
                        adicionais.append({"servico": segundo_servico, "periodo": segundo_periodo})
                    meta = registrar_metadata_apontamento(
                        conv,
                        data_apont,
                        apontado_por=engenheiro,
                        periodo_principal=periodo_principal,
                        servicos_adicionais=adicionais,
                    )
                    nova_obs = montar_observacao_operacional(turno, obs_nova, meta)
                    try:
                        supabase.table("convocacoes").update({
                            "obra_id": mapa_obras[obra_sel],
                            "status": status_sel,
                            "valor_extra": float(val_extra) if status_eh_presenca(status_sel) else 0.0,
                            "observacao": nova_obs,
                        }).eq("id", c_id).execute()
                        limpar_cache_operacional()
                        st.success(f"Apontamento de {colab.get('nome','-')} salvo.")
                        st.rerun()
                    except Exception:
                        st.error("Não foi possível salvar. Tente novamente.")


def render_indisponibilidades_admin():
    st.markdown("## 🚫 Indisponibilidade")
    st.caption("Controle manual do Paulo para férias, atestados e afastamentos. Pessoas indisponíveis ficam bloqueadas na convocação.")
    if not colaboradores:
        st.info("Nenhum colaborador cadastrado.")
        return
    labels = {f"{c.get('nome')} ({c.get('funcao','-')})": c.get("id") for c in sorted(colaboradores, key=lambda x: normalizar(x.get('nome','')))}
    c1, c2 = st.columns(2)
    with c1:
        pessoa = st.selectbox("Colaborador", list(labels.keys()), key="indisp_pessoa")
        motivo = st.selectbox("Motivo", ["Férias", "Atestado", "Afastamento", "Outro"], key="indisp_motivo")
    with c2:
        inicio = st.date_input("Início", value=agora_aproar().date(), format="DD/MM/YYYY", key="indisp_ini")
        fim = st.date_input("Fim", value=agora_aproar().date(), format="DD/MM/YYYY", key="indisp_fim")
    obs = st.text_input("Observação (opcional)", key="indisp_obs")
    if st.button("REGISTRAR INDISPONIBILIDADE", type="primary", use_container_width=True):
        ok, msg = salvar_indisponibilidade(labels[pessoa], motivo, inicio, fim, obs)
        (st.success if ok else st.error)(msg)
        if ok:
            st.rerun()

    st.markdown("### Registros")
    registros = listar_indisponibilidades()
    registros.sort(key=lambda x: (x.get("inicio", ""), x.get("fim", "")), reverse=True)
    if not registros:
        st.info("Nenhuma indisponibilidade registrada.")
        return
    for item in registros:
        colab = dict_colaboradores.get(item.get("colaborador_id"), {})
        with st.container(border=True):
            a, b = st.columns([4, 1])
            with a:
                st.markdown(f"**{colab.get('nome','Colaborador não encontrado')}** — {item.get('motivo','Indisponível')}")
                st.caption(f"{item.get('inicio','')} a {item.get('fim','')}" + (f" • {item.get('observacao')}" if item.get('observacao') else ""))
            with b:
                if st.button("Excluir", key=f"indisp_del_{item.get('id')}", use_container_width=True):
                    if excluir_indisponibilidade(item.get("id")):
                        st.rerun()
                    st.error("Não foi possível excluir.")

# --- FUNÇÕES DO PORTAL FINANCEIRO ---
def obter_ciclo_financeiro(data_ref=None):
    """Retorna o ciclo semanal de extras: terça-feira até segunda-feira."""
    data_ref = data_ref or datetime.date.today()
    dias_desde_terca = (data_ref.weekday() - 1) % 7  # terça = 1
    inicio = data_ref - datetime.timedelta(days=dias_desde_terca)
    fim = inicio + datetime.timedelta(days=6)
    pagamento = fim + datetime.timedelta(days=1)
    return inicio, fim, pagamento


def listar_ciclos_financeiros(qtd=26):
    """Gera ciclos semanais recentes para consulta do Financeiro."""
    hoje = datetime.date.today()
    inicio_atual, _, _ = obter_ciclo_financeiro(hoje)
    ciclos = []
    for i in range(qtd):
        inicio = inicio_atual - datetime.timedelta(days=7 * i)
        fim = inicio + datetime.timedelta(days=6)
        pagamento = fim + datetime.timedelta(days=1)
        em_aberto = hoje <= fim and hoje >= inicio
        ciclos.append({
            "inicio": inicio,
            "fim": fim,
            "pagamento": pagamento,
            "em_aberto": em_aberto,
            "rotulo": (
                f"{inicio.strftime('%d/%m/%Y')} a {fim.strftime('%d/%m/%Y')}"
                + (" • EM ABERTO" if em_aberto else f" • pagamento {pagamento.strftime('%d/%m/%Y')}")
            )
        })
    return ciclos


def carregar_dados_financeiro(data_inicio, data_fim):
    """Busca convocações do período e prepara extras e faltas/atestados sem expor Obra/Serviço."""
    try:
        registros = (
            supabase.table("convocacoes")
            .select("*")
            .gte("data", data_inicio.isoformat())
            .lte("data", data_fim.isoformat())
            .execute().data or []
        )
    except Exception:
        registros = []

    extras = []
    ausencias = []
    for row in registros:
        obra = dict_obras.get(row.get("obra_id"), {})
        colab = dict_colaboradores.get(row.get("colaborador_id"), {})
        unidade = str(obra.get("unidade") or "NÃO IDENTIFICADA")
        nome = str(colab.get("nome") or "NÃO IDENTIFICADO")
        funcao = str(colab.get("funcao") or "-")
        status = normalizar_status_operacional(row.get("status") or "")
        data_iso = str(row.get("data") or "")
        try:
            data_br = datetime.date.fromisoformat(data_iso).strftime("%d/%m/%Y")
        except Exception:
            data_br = data_iso
        try:
            valor_extra = float(row.get("valor_extra") or 0.0)
        except Exception:
            valor_extra = 0.0

        base = {
            "Data": data_br,
            "Data ISO": data_iso,
            "Colaborador": nome,
            "Função": funcao,
            "Unidade": unidade,
            "Engenheiro": str(row.get("engenheiro") or "N/A"),
            "Status": status,
        }

        # Financeiro considera como extra qualquer valor lançado pelo engenheiro,
        # independentemente do status do apontamento.
        if valor_extra > 0 and status_eh_presenca(status):
            item_extra = dict(base)
            item_extra["Valor Extra (R$)"] = valor_extra
            extras.append(item_extra)

        if status in ["Falta", "Atestado"]:
            ausencias.append(dict(base))

    extras.sort(key=lambda x: (x.get("Data ISO", ""), normalizar(x.get("Colaborador", ""))))
    ausencias.sort(key=lambda x: (x.get("Data ISO", ""), normalizar(x.get("Colaborador", ""))))
    return extras, ausencias


def resumir_extras_financeiro(extras):
    """Consolida as extras por colaborador para o pagamento semanal."""
    if not extras:
        return pd.DataFrame(columns=["Colaborador", "Função", "Unidades", "Lançamentos", "Total Extra (R$)"])

    df = pd.DataFrame(extras)
    resumo = (
        df.groupby(["Colaborador", "Função"], dropna=False)
        .agg(
            Unidades=("Unidade", lambda s: ", ".join(sorted(set(str(v) for v in s if str(v).strip())))),
            Lançamentos=("Valor Extra (R$)", "size"),
            **{"Total Extra (R$)": ("Valor Extra (R$)", "sum")}
        )
        .reset_index()
        .sort_values(by=["Total Extra (R$)", "Colaborador"], ascending=[False, True])
    )
    return resumo


def gerar_excel_financeiro(extras, ausencias, data_inicio, data_fim, data_pagamento):
    """Gera relatório financeiro semanal em Excel."""
    wb = openpyxl.Workbook()
    ws_resumo = wb.active
    ws_resumo.title = "Resumo Extras"

    fill_titulo = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    fill_header = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    font_titulo = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    font_header = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    borda = Border(
        left=Side(style="thin", color="CBD5E1"), right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"), bottom=Side(style="thin", color="CBD5E1")
    )

    def cabecalho_planilha(ws, titulo, subtitulo, total_colunas):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_colunas)
        c = ws.cell(1, 1, titulo)
        c.font = font_titulo
        c.fill = fill_titulo
        c.alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_colunas)
        ws.cell(2, 1, subtitulo).font = Font(name="Arial", size=9, italic=True, color="64748B")

    resumo = resumir_extras_financeiro(extras)
    total_extra = sum(float(x.get("Valor Extra (R$)") or 0) for x in extras)
    subtitulo = (
        f"Ciclo: {data_inicio.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')} | "
        f"Pagamento: {data_pagamento.strftime('%d/%m/%Y')} | Total: {formatar_reais(total_extra)}"
    )
    cabecalho_planilha(ws_resumo, "APROAR - RELATÓRIO SEMANAL DE EXTRAS", subtitulo, 5)
    headers_resumo = ["Colaborador", "Função", "Unidades", "Lançamentos", "Total Extra (R$)"]
    for ci, nome in enumerate(headers_resumo, 1):
        cell = ws_resumo.cell(4, ci, nome)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center")
    for ri, (_, r) in enumerate(resumo.iterrows(), 5):
        vals = [r["Colaborador"], r["Função"], r["Unidades"], int(r["Lançamentos"]), float(r["Total Extra (R$)"])]
        for ci, val in enumerate(vals, 1):
            cell = ws_resumo.cell(ri, ci, val)
            cell.border = borda
            cell.font = Font(name="Arial", size=9)
            if ci == 5:
                cell.number_format = 'R$ #,##0.00'
    ws_resumo.freeze_panes = "A5"
    for col, largura in {"A": 38, "B": 28, "C": 38, "D": 14, "E": 20}.items():
        ws_resumo.column_dimensions[col].width = largura

    ws_det = wb.create_sheet("Detalhe Extras")
    cabecalho_planilha(ws_det, "DETALHAMENTO DE EXTRAS", subtitulo, 6)
    headers_det = ["Data", "Colaborador", "Função", "Unidade", "Engenheiro", "Valor Extra (R$)"]
    for ci, nome in enumerate(headers_det, 1):
        cell = ws_det.cell(4, ci, nome)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center")
    for ri, item in enumerate(extras, 5):
        vals = [item["Data"], item["Colaborador"], item["Função"], item["Unidade"], item["Engenheiro"], item["Valor Extra (R$)"]]
        for ci, val in enumerate(vals, 1):
            cell = ws_det.cell(ri, ci, val)
            cell.border = borda
            cell.font = Font(name="Arial", size=9)
            if ci == 6:
                cell.number_format = 'R$ #,##0.00'
    ws_det.freeze_panes = "A5"
    for col, largura in {"A": 14, "B": 38, "C": 28, "D": 28, "E": 18, "F": 20}.items():
        ws_det.column_dimensions[col].width = largura

    ws_aus = wb.create_sheet("Faltas e Atestados")
    cabecalho_planilha(
        ws_aus,
        "FALTAS E ATESTADOS DO CICLO",
        f"Período: {data_inicio.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')}",
        6
    )
    headers_aus = ["Data", "Colaborador", "Função", "Unidade", "Status", "Engenheiro"]
    for ci, nome in enumerate(headers_aus, 1):
        cell = ws_aus.cell(4, ci, nome)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center")
    for ri, item in enumerate(ausencias, 5):
        vals = [item["Data"], item["Colaborador"], item["Função"], item["Unidade"], item["Status"], item["Engenheiro"]]
        for ci, val in enumerate(vals, 1):
            cell = ws_aus.cell(ri, ci, val)
            cell.border = borda
            cell.font = Font(name="Arial", size=9)
    ws_aus.freeze_panes = "A5"
    for col, largura in {"A": 14, "B": 38, "C": 28, "D": 28, "E": 16, "F": 18}.items():
        ws_aus.column_dimensions[col].width = largura

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def gerar_pdf_financeiro(extras, ausencias, data_inicio, data_fim, data_pagamento):
    """Gera um relatório financeiro compacto em PDF, sem Obra/Serviço."""
    pdf = FPDF(orientation="L")
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.add_page()
    pdf.set_font("Arial", "B", 14)
    pdf.cell(0, 9, to_latin("APROAR - RELATÓRIO FINANCEIRO SEMANAL"), ln=True, align="C")
    pdf.set_font("Arial", "", 9)
    pdf.cell(
        0, 7,
        to_latin(
            f"Ciclo de extras: {data_inicio.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')} | "
            f"Pagamento previsto: {data_pagamento.strftime('%d/%m/%Y')}"
        ),
        ln=True, align="C"
    )
    pdf.ln(3)

    resumo = resumir_extras_financeiro(extras)
    total_extra = sum(float(x.get("Valor Extra (R$)") or 0) for x in extras)
    pdf.set_font("Arial", "B", 11)
    pdf.cell(0, 7, to_latin(f"EXTRAS - TOTAL A PAGAR: {formatar_reais(total_extra)}"), ln=True)

    widths = [72, 48, 72, 28, 36]
    headers = ["Colaborador", "Função", "Unidade(s)", "Lanç.", "Total"]
    pdf.set_font("Arial", "B", 8)
    for w, h in zip(widths, headers):
        pdf.cell(w, 6, to_latin(h), border=1, align="C")
    pdf.ln()
    pdf.set_font("Arial", "", 8)
    if resumo.empty:
        pdf.cell(sum(widths), 6, to_latin("Nenhuma extra lançada neste ciclo."), border=1, ln=True)
    else:
        for _, r in resumo.iterrows():
            vals = [
                str(r["Colaborador"])[:34], str(r["Função"])[:22], str(r["Unidades"])[:33],
                str(int(r["Lançamentos"])), formatar_reais(float(r["Total Extra (R$)"]))
            ]
            aligns = ["L", "L", "L", "C", "R"]
            for w, v, a in zip(widths, vals, aligns):
                pdf.cell(w, 6, to_latin(v), border=1, align=a)
            pdf.ln()

    pdf.ln(5)
    pdf.set_font("Arial", "B", 11)
    pdf.cell(0, 7, to_latin("FALTAS E ATESTADOS"), ln=True)
    widths2 = [25, 68, 45, 45, 30, 40]
    headers2 = ["Data", "Colaborador", "Função", "Unidade", "Status", "Engenheiro"]
    pdf.set_font("Arial", "B", 8)
    for w, h in zip(widths2, headers2):
        pdf.cell(w, 6, to_latin(h), border=1, align="C")
    pdf.ln()
    pdf.set_font("Arial", "", 8)
    if not ausencias:
        pdf.cell(sum(widths2), 6, to_latin("Nenhuma falta ou atestado neste ciclo."), border=1, ln=True)
    else:
        for item in ausencias:
            vals = [
                item["Data"], item["Colaborador"][:32], item["Função"][:20],
                item["Unidade"][:20], item["Status"], item["Engenheiro"][:18]
            ]
            aligns = ["C", "L", "L", "L", "C", "C"]
            for w, v, a in zip(widths2, vals, aligns):
                pdf.cell(w, 6, to_latin(v), border=1, align=a)
            pdf.ln()

    return pdf.output(dest="S").encode("latin1")

# --- ACESSO SIMPLES: VISUALIZAÇÃO LIVRE + UMA SENHA ÚNICA DE EDIÇÃO ---
def _obter_senha_edicao():
    """
    Aceita qualquer um destes formatos no secrets.toml:

    SENHA_EDICAO = "sua_senha"

    ou:

    [acesso]
    senha_edicao = "sua_senha"

    Quem não informar a senha permanece em modo somente leitura.
    """
    senha = ""
    try:
        senha = str(st.secrets.get("SENHA_EDICAO", "") or "").strip()
    except Exception:
        senha = ""

    if senha:
        return senha

    try:
        bloco = st.secrets.get("acesso", {})
        senha = str(bloco.get("senha_edicao", "") or "").strip()
    except Exception:
        senha = ""
    return senha


def _edicao_liberada():
    return bool(st.session_state.get("edicao_liberada", False))


def _render_desbloqueio_edicao():
    """Mostra um acesso discreto para quem possui a senha única de edição."""
    senha_configurada = _obter_senha_edicao()
    with st.expander("🔐 Tenho acesso para editar", expanded=False):
        if not senha_configurada:
            st.info("A senha de edição ainda não foi configurada nos Secrets. O sistema está em modo somente leitura.")
            st.code('SENHA_EDICAO = "sua_senha"', language="toml")
            return

        st.caption("Digite a senha de edição. Não é necessário usuário individual.")
        with st.form("form_desbloquear_edicao", clear_on_submit=True):
            senha_digitada = st.text_input("Senha de edição", type="password")
            liberar = st.form_submit_button("LIBERAR EDIÇÃO", type="primary", use_container_width=True)

        if liberar:
            if hmac.compare_digest(str(senha_digitada), str(senha_configurada)):
                st.session_state["edicao_liberada"] = True
                st.success("Edição liberada.")
                st.rerun()
            else:
                st.error("Senha incorreta.")


def _logout_disponivel():
    """No novo modelo, 'sair' significa voltar ao modo somente leitura."""
    if _edicao_liberada():
        if st.button("🔒 BLOQUEAR EDIÇÃO", key="bloquear_edicao_topo"):
            st.session_state["edicao_liberada"] = False
            st.rerun()

parametros_url = st.query_params
modo_campo_solicitado = "eng" in parametros_url or parametros_url.get("modo") in ["campo", "eng"]
modo_financeiro_solicitado = (
    "financeiro" in parametros_url
    or "fin" in parametros_url
    or parametros_url.get("modo") in ["financeiro", "fin"]
)
modo_visualizador_solicitado = "view" in parametros_url or parametros_url.get("modo") in ["visualizador", "view"]

edicao_liberada = _edicao_liberada()
if not edicao_liberada:
    modo_campo = False
    modo_financeiro = False
    modo_visualizador = True
else:
    modo_campo = modo_campo_solicitado
    modo_financeiro = modo_financeiro_solicitado
    modo_visualizador = modo_visualizador_solicitado

if modo_visualizador:
    st.markdown("## 👁️ Portal de consulta")
    st.caption("Sem a senha, o sistema fica somente para consulta. Dashboard, Relatórios e Indicadores permanecem disponíveis.")
    _render_desbloqueio_edicao()
    secao_view = st.radio(
        "Navegação",
        ["🎛️ DASHBOARD", "📊 RELATÓRIOS", "📈 INDICADORES"],
        horizontal=True,
        label_visibility="collapsed",
        key="nav_visualizador_geral",
    )
    if secao_view == "🎛️ DASHBOARD":
        render_dashboard_consulta("view_dash")
    elif secao_view == "📊 RELATÓRIOS":
        render_relatorio_visualizador("view_rel")
    else:
        render_indicadores_cumprimento("view_ind")

elif modo_campo:
    # ==========================================
    # PORTAL DO ENGENHEIRO — FLUXO OPERACIONAL SIMPLIFICADO
    # ==========================================
    st.markdown("## 👷 Portal do Engenheiro")
    st.caption("Planeje a equipe de amanhã, faça o apontamento de hoje e consulte disponibilidade sem sair desta tela.")
    _logout_disponivel()

    def _buscar_convocacoes_campo(engenheiro, data_ref):
        try:
            return (
                supabase.table("convocacoes")
                .select("*")
                .eq("engenheiro", engenheiro)
                .eq("data", data_ref.isoformat())
                .execute().data or []
            )
        except Exception:
            return []

    def _enriquecer_convocacoes_campo(registros):
        saida = []
        for registro in registros or []:
            item = dict(registro)
            item["dados_obra"] = dict_obras.get(
                item.get("obra_id"),
                {"unidade": "Desconhecida", "nome": NOME_OBRA_PLACEHOLDER},
            )
            saida.append(item)
        return saida

    def _convocacao_apontada_campo(conv):
        obra = conv.get("dados_obra") or dict_obras.get(conv.get("obra_id"), {})
        return bool(obra) and not eh_obra_placeholder(obra)

    def _mudar_secao_campo(secao):
        st.session_state["campo_secao"] = secao

    hoje_campo = datetime.date.today()
    amanha_campo = proximo_dia_util(hoje_campo)

    topo_c1, topo_c2 = st.columns([1.2, 2.8])
    with topo_c1:
        engenheiro_campo = st.selectbox(
            "Engenheiro",
            ENGENHEIROS,
            key="engenheiro_campo_global",
        )
    with topo_c2:
        st.markdown(
            f"<div style='padding-top:29px;color:#94A3B8;font-size:13px;'>"
            f"Hoje: <b style='color:#F8FAFC'>{hoje_campo.strftime('%d/%m/%Y')}</b> &nbsp;•&nbsp; "
            f"Próximo dia: <b style='color:#F8FAFC'>{amanha_campo.strftime('%d/%m/%Y')}</b>"
            f"</div>",
            unsafe_allow_html=True,
        )

    secoes_principais_campo = [
        "📌 RESUMO", "👥 CONVOCADOS", "👥 DISPONIBILIDADE",
        "🎛️ DASHBOARD", "📊 RELATÓRIOS", "📈 INDICADORES"
    ]
    secoes_ocultas_campo = ["✅ APONTAMENTO", "📋 EQUIPE DE AMANHÃ"]
    secoes_validas_campo = secoes_principais_campo + secoes_ocultas_campo

    if st.session_state.get("campo_secao") not in secoes_validas_campo:
        st.session_state["campo_secao"] = "📌 RESUMO"

    secao_campo = st.session_state["campo_secao"]
    secao_nav = st.radio(
        "Navegação",
        secoes_principais_campo,
        index=secoes_principais_campo.index(secao_campo) if secao_campo in secoes_principais_campo else 0,
        horizontal=True,
        label_visibility="collapsed",
        key="campo_nav_principal_melhorias",
    )
    if secao_campo in secoes_principais_campo and secao_nav != secao_campo:
        st.session_state["campo_secao"] = secao_nav
        st.rerun()

    # --------------------------------------------------------------
    # RESUMO DO ENGENHEIRO
    # --------------------------------------------------------------
    if secao_campo == "📌 RESUMO":
        convocacoes_hoje_campo = _enriquecer_convocacoes_campo(
            _buscar_convocacoes_campo(engenheiro_campo, hoje_campo)
        )
        convocacoes_amanha_campo = _enriquecer_convocacoes_campo(
            _buscar_convocacoes_campo(engenheiro_campo, amanha_campo)
        )

        total_hoje = len(convocacoes_hoje_campo)
        apontados_hoje = sum(1 for c in convocacoes_hoje_campo if _convocacao_apontada_campo(c))
        pendentes_hoje = max(0, total_hoje - apontados_hoje)
        faltas_hoje = sum(1 for c in convocacoes_hoje_campo if str(c.get("status")) == "Falta")
        atestados_hoje = sum(1 for c in convocacoes_hoje_campo if str(c.get("status")) == "Atestado")

        st.markdown("### Resumo do dia")
        k1, k2 = st.columns(2)
        k1.metric("Equipe", total_hoje)
        k2.metric("Apontados", apontados_hoje)

        k3, k4 = st.columns(2)
        k3.metric("Pendentes", pendentes_hoje)
        k4.metric("Faltas / Atestados", faltas_hoje + atestados_hoje)

        if pendentes_hoje:
            st.warning(f"⚠️ Você ainda tem **{pendentes_hoje} apontamento(s)** para concluir hoje.")
        elif total_hoje:
            st.success("✅ Todos os colaboradores de hoje já têm Obra/Serviço definida no apontamento.")
        else:
            st.info("Nenhuma equipe foi convocada para você hoje.")

        st.button(
            "✅ FAZER APONTAMENTO DE HOJE",
            type="primary",
            use_container_width=True,
            on_click=_mudar_secao_campo,
            args=("✅ APONTAMENTO",),
            key="btn_ir_apontamento_resumo_campo",
        )
        st.caption("Defina obra/serviço e ajuste presenças.")

        st.button(
            "📋 MONTAR EQUIPE DO PRÓXIMO DIA",
            use_container_width=True,
            on_click=_mudar_secao_campo,
            args=("📋 EQUIPE DE AMANHÃ",),
            key="btn_ir_convocacao_resumo_campo",
        )
        st.caption("Escolha unidade, turno e colaboradores.")

        st.markdown("### Próximo dia")
        if convocacoes_amanha_campo:
            unidades_amanha = sorted({str((c.get("dados_obra") or {}).get("unidade") or "-") for c in convocacoes_amanha_campo})
            st.info(
                f"**{len(convocacoes_amanha_campo)} pessoa(s) convocada(s)** para {amanha_campo.strftime('%d/%m/%Y')}"
                + (f" • {', '.join(unidades_amanha)}" if unidades_amanha else "")
            )
            with st.expander("Ver equipe já convocada", expanded=False):
                for conv in convocacoes_amanha_campo:
                    colab = dict_colaboradores.get(conv.get("colaborador_id"), {})
                    unidade = (conv.get("dados_obra") or {}).get("unidade", "-")
                    turno, _ = decompor_observacao_operacional(conv.get("observacao") or "")
                    st.markdown(f"• **{colab.get('nome', 'Não identificado')}** — {unidade} • {turno}")
        else:
            st.caption("Nenhuma convocação sua registrada para o próximo dia.")

    # --------------------------------------------------------------
    # CONVOCADOS PELO ENGENHEIRO
    # --------------------------------------------------------------
    elif secao_campo == "👥 CONVOCADOS":
        convocacoes_hoje_campo = _enriquecer_convocacoes_campo(
            _buscar_convocacoes_campo(engenheiro_campo, hoje_campo)
        )
        convocacoes_amanha_campo = _enriquecer_convocacoes_campo(
            _buscar_convocacoes_campo(engenheiro_campo, amanha_campo)
        )

        st.markdown("### 👥 Convocados por você")
        st.caption("Veja rapidamente quem foi convocado por você hoje e no próximo dia.")

        c1, c2 = st.columns(2)
        c1.metric("Hoje", len(convocacoes_hoje_campo))
        c2.metric("Próximo dia", len(convocacoes_amanha_campo))

        st.markdown(f"#### Hoje • {hoje_campo.strftime('%d/%m/%Y')}")
        if convocacoes_hoje_campo:
            for conv in convocacoes_hoje_campo:
                colab = dict_colaboradores.get(conv.get("colaborador_id"), {})
                unidade = (conv.get("dados_obra") or {}).get("unidade", "-")
                turno, _ = decompor_observacao_operacional(conv.get("observacao") or "")
                status = str(conv.get("status") or "Pendente")
                st.markdown(
                    f"• **{colab.get('nome', 'Não identificado')}** — {unidade} • {turno} • {status}"
                )
        else:
            st.caption("Nenhuma pessoa convocada por você hoje.")

        st.markdown(f"#### Próximo dia • {amanha_campo.strftime('%d/%m/%Y')}")
        if convocacoes_amanha_campo:
            for conv in convocacoes_amanha_campo:
                colab = dict_colaboradores.get(conv.get("colaborador_id"), {})
                unidade = (conv.get("dados_obra") or {}).get("unidade", "-")
                turno, _ = decompor_observacao_operacional(conv.get("observacao") or "")
                st.markdown(
                    f"• **{colab.get('nome', 'Não identificado')}** — {unidade} • {turno}"
                )
        else:
            st.caption("Nenhuma pessoa convocada por você para o próximo dia.")

    # --------------------------------------------------------------
    # APONTAMENTO — MULTISSERVIÇO / RETROATIVO / EXTRA SEPARADA
    # --------------------------------------------------------------
    elif secao_campo == "✅ APONTAMENTO":
        if st.button("⬅️ Voltar ao resumo", key="voltar_resumo_apont_campo_melhorias"):
            _mudar_secao_campo("📌 RESUMO")
            st.rerun()
        render_apontamento_operacional(engenheiro_campo, "campo_apont_melhorias")

    # --------------------------------------------------------------
    # CONVOCAÇÃO — PRÓXIMO DIA
    # --------------------------------------------------------------
    elif secao_campo == "📋 EQUIPE DE AMANHÃ":
        if st.button("⬅️ Voltar ao resumo", key="voltar_resumo_conv_campo"):
            _mudar_secao_campo("📌 RESUMO")
            st.rerun()
        st.markdown("### 📋 Equipe do próximo dia")
        st.caption("Escolha a Unidade e as pessoas. A Obra/Serviço específica continua sendo definida no apontamento.")

        data_conv_auto = amanha_campo
        ja_convocados = _enriquecer_convocacoes_campo(_buscar_convocacoes_campo(engenheiro_campo, data_conv_auto))

        hc1, hc2 = st.columns(2)
        hc1.info(f"📅 **Data:** {data_conv_auto.strftime('%d/%m/%Y')}")
        hc2.info(f"👥 **Já convocados por você:** {len(ja_convocados)}")

        if ja_convocados:
            with st.expander("Ver equipe já convocada", expanded=False):
                for conv in ja_convocados:
                    colab = dict_colaboradores.get(conv.get("colaborador_id"), {})
                    unidade = (conv.get("dados_obra") or {}).get("unidade", "-")
                    turno, _ = decompor_observacao_operacional(conv.get("observacao") or "")
                    st.markdown(f"• **{colab.get('nome', 'Não identificado')}** — {unidade} • {turno}")

        if not obras:
            st.info("Cadastre pelo menos uma Unidade/Obra na administração.")
        else:
            cc1, cc2 = st.columns(2)
            unidades_unicas = UNIDADES_APROAR.copy()
            with cc1:
                unidade_selecionada = st.selectbox("Unidade", unidades_unicas, key="u_c_sel_novo")
            with cc2:
                turno_conv_campo = st.selectbox("Turno", ["Integral", "Manhã", "Tarde", "Noite"], key="turno_conv_campo_novo")

            funcoes_disponiveis = sorted({c.get("funcao", "") for c in colaboradores if c.get("funcao")})
            filtro_funcao = st.selectbox("Filtrar por função", ["TODAS"] + funcoes_disponiveis, key="f_c_sel_campo_novo")
            colabs_filtrados = [c for c in colaboradores if filtro_funcao == "TODAS" or c.get("funcao") == filtro_funcao]

            mapa_colab_opcoes = {f"{c['nome']}  ({c.get('funcao', '-')})": c["id"] for c in colabs_filtrados}
            equipe_selecionada = st.multiselect(
                "Selecionar colaboradores",
                list(mapa_colab_opcoes.keys()),
                placeholder="Digite para buscar pelo nome...",
                key="eq_c_sel_campo_novo",
            )

            avulso_campo = False
            nome_manual_campo = ""
            tipo_manual_campo = "Profissional"
            funcao_manual_campo = ""
            with st.expander("➕ Adicionar nome que não está na lista", expanded=False):
                avulso_campo = st.checkbox("É avulso?", key="avulso_conv_campo_novo")
                nome_manual_campo = st.text_input(
                    "Nome",
                    placeholder="Digite o nome completo...",
                    key="nome_manual_conv_campo_novo",
                )
                if avulso_campo or nome_manual_campo.strip():
                    cm1, cm2 = st.columns(2)
                    with cm1:
                        tipo_manual_campo = st.selectbox(
                            "Categoria da diária", ["Profissional", "Ajudante"], key="tipo_manual_conv_campo_novo"
                        )
                    with cm2:
                        funcao_manual_campo = st.text_input(
                            "Função (opcional)", placeholder="Ex.: pintor, eletricista...", key="funcao_manual_conv_campo_novo"
                        )

            if st.button("CONFIRMAR CONVOCAÇÃO", type="primary", use_container_width=True, key="btn_conv_campo_novo"):
                if avulso_campo and not nome_manual_campo.strip():
                    st.warning("Para convocar como avulso, digite o nome do colaborador.")
                elif not equipe_selecionada and not nome_manual_campo.strip():
                    st.warning("Selecione pelo menos um colaborador ou digite um nome.")
                else:
                    obra_id_placeholder = obter_obra_placeholder_unidade(unidade_selecionada)
                    if not obra_id_placeholder:
                        st.error(
                            f"Não foi possível preparar a Unidade **{unidade_selecionada}** "
                            "para a convocação."
                        )
                        detalhe_placeholder = st.session_state.get("erro_placeholder_unidade", "")
                        if detalhe_placeholder:
                            with st.expander("Diagnóstico técnico"):
                                st.code(detalhe_placeholder)
                    else:
                        pessoas = []
                        for label_colab in equipe_selecionada:
                            c_id = mapa_colab_opcoes[label_colab]
                            nome_existente = dict_colaboradores.get(c_id, {}).get("nome", label_colab.split("  (")[0])
                            pessoas.append((c_id, nome_existente))

                        if nome_manual_campo.strip():
                            c_id_manual, colab_manual, msg_manual = criar_ou_obter_colaborador_manual(
                                nome_manual_campo, tipo_manual_campo, funcao_manual_campo, avulso=avulso_campo
                            )
                            if c_id_manual:
                                pessoas.append((c_id_manual, colab_manual.get("nome", nome_manual_campo)))
                            else:
                                st.error(msg_manual)

                        pessoas_unicas = []
                        ids_vistos = set()
                        for cid, nome_pessoa in pessoas:
                            if cid and cid not in ids_vistos:
                                pessoas_unicas.append((cid, nome_pessoa))
                                ids_vistos.add(cid)

                        sucessos = 0
                        avisos = []
                        for c_id, nome_pessoa in pessoas_unicas:
                            ok, motivo = inserir_convocacao_segura(
                                obra_id_placeholder, c_id, data_conv_auto, engenheiro_campo, turno_conv_campo
                            )
                            if ok:
                                sucessos += 1
                            else:
                                avisos.append(f"{nome_pessoa}: {motivo}.")

                        if sucessos:
                            limpar_cache_operacional()
                            st.success(
                                f"✅ {sucessos} colaborador(es) convocado(s) para {unidade_selecionada} "
                                f"em {data_conv_auto.strftime('%d/%m/%Y')} • Turno {turno_conv_campo}."
                            )
                        for aviso in avisos:
                            st.warning(aviso)

    # --------------------------------------------------------------
    # DISPONIBILIDADE
    # --------------------------------------------------------------
    elif secao_campo == "🎛️ DASHBOARD":
        render_dashboard_consulta("campo_dash_melhorias", engenheiro_campo)

    elif secao_campo == "📊 RELATÓRIOS":
        render_relatorio_visualizador("campo_rel_melhorias", engenheiro_campo)

    elif secao_campo == "📈 INDICADORES":
        render_indicadores_cumprimento("campo_ind_melhorias", engenheiro_campo)

    elif secao_campo == "👥 DISPONIBILIDADE":
        render_aba_disponibilidade("campo_novo")

elif modo_financeiro:
    # ==========================================
    # PORTAL FINANCEIRO (?financeiro)
    # ==========================================
    st.markdown("### 💰 ACESSO FINANCEIRO")
    st.caption("Conferência semanal de extras, faltas e atestados. As extras são fechadas em ciclos de terça-feira a segunda-feira.")
    _logout_disponivel()

    ciclos_fin = listar_ciclos_financeiros(26)
    mapa_ciclos_fin = {c["rotulo"]: c for c in ciclos_fin}

    # Na terça-feira, o financeiro normalmente paga o ciclo que encerrou na segunda anterior.
    indice_padrao_fin = 1 if datetime.date.today().weekday() == 1 and len(ciclos_fin) > 1 else 0
    ciclo_rotulo_fin = st.selectbox(
        "Ciclo semanal:",
        list(mapa_ciclos_fin.keys()),
        index=indice_padrao_fin,
        key="ciclo_financeiro"
    )
    ciclo_fin = mapa_ciclos_fin[ciclo_rotulo_fin]
    data_ini_fin = ciclo_fin["inicio"]
    data_fim_fin = ciclo_fin["fim"]
    data_pag_fin = ciclo_fin["pagamento"]

    cf1, cf2, cf3 = st.columns(3)
    cf1.metric("INÍCIO", data_ini_fin.strftime("%d/%m/%Y"))
    cf2.metric("FIM", data_fim_fin.strftime("%d/%m/%Y"))
    cf3.metric("PAGAMENTO", data_pag_fin.strftime("%d/%m/%Y"))

    extras_fin, ausencias_fin = carregar_dados_financeiro(data_ini_fin, data_fim_fin)
    total_extra_fin = sum(float(x.get("Valor Extra (R$)") or 0.0) for x in extras_fin)
    nomes_extra_fin = {normalizar(x.get("Colaborador", "")) for x in extras_fin}
    total_faltas_fin = sum(1 for x in ausencias_fin if x.get("Status") == "Falta")
    total_atest_fin = sum(1 for x in ausencias_fin if x.get("Status") == "Atestado")

    tab_fin_extra, tab_fin_aus, tab_fin_rel = st.tabs([
        "💸 EXTRAS", "🚫 FALTAS / ATESTADOS", "📄 RELATÓRIO"
    ])

    with tab_fin_extra:
        fm1, fm2, fm3 = st.columns(3)
        fm1.metric("TOTAL A PAGAR", formatar_reais(total_extra_fin))
        fm2.metric("COLABORADORES", len(nomes_extra_fin))
        fm3.metric("LANÇAMENTOS", len(extras_fin))

        st.markdown("### Consolidado por colaborador")
        resumo_fin = resumir_extras_financeiro(extras_fin)
        if resumo_fin.empty:
            st.info("Nenhuma extra foi lançada neste ciclo.")
        else:
            resumo_view = resumo_fin.copy()
            resumo_view["Total Extra"] = resumo_view["Total Extra (R$)"].apply(formatar_reais)
            resumo_view = resumo_view.drop(columns=["Total Extra (R$)"])
            st.dataframe(resumo_view, use_container_width=True, hide_index=True)

            st.markdown("### Detalhamento por dia")
            detalhe_extra_view = pd.DataFrame(extras_fin)[[
                "Data", "Colaborador", "Função", "Unidade", "Engenheiro", "Valor Extra (R$)"
            ]].copy()
            detalhe_extra_view["Valor Extra"] = detalhe_extra_view["Valor Extra (R$)"].apply(formatar_reais)
            detalhe_extra_view = detalhe_extra_view.drop(columns=["Valor Extra (R$)"])
            st.dataframe(detalhe_extra_view, use_container_width=True, hide_index=True)

    with tab_fin_aus:
        fa1, fa2, fa3 = st.columns(3)
        fa1.metric("FALTAS", total_faltas_fin)
        fa2.metric("ATESTADOS", total_atest_fin)
        fa3.metric("TOTAL OCORRÊNCIAS", len(ausencias_fin))

        if not ausencias_fin:
            st.info("Nenhuma falta ou atestado foi registrado neste ciclo.")
        else:
            df_aus_fin = pd.DataFrame(ausencias_fin)[[
                "Data", "Colaborador", "Função", "Unidade", "Status", "Engenheiro"
            ]]
            st.dataframe(df_aus_fin, use_container_width=True, hide_index=True)

            st.markdown("### Resumo nominal")
            resumo_aus_fin = (
                df_aus_fin.groupby(["Colaborador", "Função"], dropna=False)
                .agg(
                    Faltas=("Status", lambda s: int((s == "Falta").sum())),
                    Atestados=("Status", lambda s: int((s == "Atestado").sum())),
                    Unidades=("Unidade", lambda s: ", ".join(sorted(set(str(v) for v in s if str(v).strip()))))
                )
                .reset_index()
            )
            resumo_aus_fin["Total"] = resumo_aus_fin["Faltas"] + resumo_aus_fin["Atestados"]
            resumo_aus_fin = resumo_aus_fin.sort_values(by=["Total", "Colaborador"], ascending=[False, True])
            st.dataframe(resumo_aus_fin, use_container_width=True, hide_index=True)

    with tab_fin_rel:
        st.markdown("### Relatório do ciclo")
        st.write(
            f"Período **{data_ini_fin.strftime('%d/%m/%Y')} a {data_fim_fin.strftime('%d/%m/%Y')}** • "
            f"Pagamento das extras em **{data_pag_fin.strftime('%d/%m/%Y')}**."
        )
        st.info(
            f"Total de extras: {formatar_reais(total_extra_fin)} • "
            f"Faltas: {total_faltas_fin} • Atestados: {total_atest_fin}"
        )

        excel_fin = gerar_excel_financeiro(extras_fin, ausencias_fin, data_ini_fin, data_fim_fin, data_pag_fin)
        pdf_fin = gerar_pdf_financeiro(extras_fin, ausencias_fin, data_ini_fin, data_fim_fin, data_pag_fin)

        fr1, fr2 = st.columns(2)
        with fr1:
            st.download_button(
                "📊 BAIXAR RELATÓRIO EXCEL",
                data=excel_fin,
                file_name=f"financeiro_extras_{data_ini_fin.strftime('%d-%m-%Y')}_a_{data_fim_fin.strftime('%d-%m-%Y')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="download_fin_excel"
            )
        with fr2:
            st.download_button(
                "📄 BAIXAR RELATÓRIO PDF",
                data=pdf_fin,
                file_name=f"financeiro_extras_{data_ini_fin.strftime('%d-%m-%Y')}_a_{data_fim_fin.strftime('%d-%m-%Y')}.pdf",
                mime="application/pdf",
                use_container_width=True,
                key="download_fin_pdf"
            )

else:
    # ==========================================
    # PAINEL ADMINISTRATIVO (TEMA ESCURO)
    # ==========================================
    
    if "menu_ativo" not in st.session_state:
        st.session_state.menu_ativo = "🏠 INÍCIO"

    def _ir_menu_admin(destino):
        st.session_state["menu_ativo"] = destino

    with st.sidebar:
        if os.path.exists("logo.png"):
            st.image("logo.png", use_container_width=True)
        else:
            st.markdown("<h2 style='text-align: center; color: #FFFFFF; letter-spacing: 2px;'>APROAR</h2>", unsafe_allow_html=True)

        st.markdown("<p style='text-align: center; font-size: 10px; color: #94A3B8; letter-spacing: 1.5px; margin-top: -5px; margin-bottom: 16px; font-weight: 700;'>GESTÃO DE EQUIPES</p>", unsafe_allow_html=True)

        st.button("🏠 INÍCIO", key="btn_nav_inicio", use_container_width=True, on_click=_ir_menu_admin, args=("🏠 INÍCIO",))

        st.markdown("<div class='aproar-sidebar-section'>🛠 OPERAÇÃO</div>", unsafe_allow_html=True)
        for item in ["📋 CONVOCAÇÃO", "✅ APONTAMENTO", "💬 WHATSAPP", "👥 DISPONIBILIDADE", "🚫 INDISPONIBILIDADE"]:
            st.button(item, key=f"btn_nav_{item}_novo", use_container_width=True, on_click=_ir_menu_admin, args=(item,))

        st.markdown("<div class='aproar-sidebar-section'>📊 ANÁLISE E FECHAMENTO</div>", unsafe_allow_html=True)
        for item in ["🎛️ DASHBOARD", "📊 RELATÓRIOS", "📈 INDICADORES"]:
            st.button(item, key=f"btn_nav_{item}_novo", use_container_width=True, on_click=_ir_menu_admin, args=(item,))

        st.markdown("<div class='aproar-sidebar-section'>⚙ SISTEMA</div>", unsafe_allow_html=True)
        st.button("⚙️ CONFIGURAÇÕES", key="btn_nav_config_novo", use_container_width=True, on_click=_ir_menu_admin, args=("⚙️ CONFIGURAÇÕES",))

        st.markdown("---")
        st.caption("✏️ Modo edição liberado")
        if st.button("🔒 BLOQUEAR EDIÇÃO", key="bloquear_edicao_sidebar", use_container_width=True):
            st.session_state["edicao_liberada"] = False
            st.rerun()
        st.caption("APROAR Engenharia © 2026")

    menu_escolhido = st.session_state.menu_ativo

    # --- HOME ADMINISTRATIVA: O QUE PRECISA DE ATENÇÃO ---
    if menu_escolhido == "🏠 INÍCIO":
        st.markdown("## 🏠 Visão do dia")
        st.caption("A tela inicial mostra primeiro o que ainda precisa ser resolvido. As consultas detalhadas continuam nos módulos do menu.")

        hoje_admin = datetime.date.today()
        amanha_admin = proximo_dia_util(hoje_admin)
        try:
            conv_hoje_admin = supabase.table("convocacoes").select("*").eq("data", hoje_admin.isoformat()).execute().data or []
        except Exception:
            conv_hoje_admin = []
        try:
            conv_amanha_admin = supabase.table("convocacoes").select("*").eq("data", amanha_admin.isoformat()).execute().data or []
        except Exception:
            conv_amanha_admin = []

        pendentes_admin = []
        for conv in conv_hoje_admin:
            obra_conv = dict_obras.get(conv.get("obra_id"), {})
            if not obra_conv or eh_obra_placeholder(obra_conv):
                pendentes_admin.append(conv)

        total_hoje_admin = len(conv_hoje_admin)
        apontados_admin = max(0, total_hoje_admin - len(pendentes_admin))
        faltas_admin = sum(1 for c in conv_hoje_admin if normalizar_status_operacional(c.get("status")) == "Falta")
        atestados_admin = sum(1 for c in conv_hoje_admin if normalizar_status_operacional(c.get("status")) == "Atestado")

        a1, a2, a3, a4, a5 = st.columns(5)
        a1.metric("EQUIPE HOJE", total_hoje_admin)
        a2.metric("APONTADOS", apontados_admin)
        a3.metric("PENDENTES", len(pendentes_admin))
        a4.metric("FALTAS / ATESTADOS", faltas_admin + atestados_admin)
        a5.metric("CONVOCADOS AMANHÃ", len(conv_amanha_admin))

        st.markdown("### ⚠️ Precisa de atenção")
        if pendentes_admin:
            resumo_pend = {}
            for conv in pendentes_admin:
                eng = str(conv.get("engenheiro") or "NÃO INFORMADO")
                resumo_pend[eng] = resumo_pend.get(eng, 0) + 1
            df_pend = pd.DataFrame([
                {"Engenheiro": eng, "Apontamentos pendentes": qtd}
                for eng, qtd in sorted(resumo_pend.items(), key=lambda x: (-x[1], x[0]))
            ])
            st.warning(f"Existem **{len(pendentes_admin)} colaborador(es)** de hoje ainda sem Obra/Serviço definida.")
            st.dataframe(df_pend, use_container_width=True, hide_index=True)
        elif total_hoje_admin:
            st.success("✅ Não há apontamentos pendentes de hoje.")
        else:
            st.info("Nenhuma convocação registrada para hoje.")


        # Conflitos de convocação registrados para conferência do Paulo
        try:
            desde_conflitos = (hoje_admin - datetime.timedelta(days=45)).isoformat()
            regs_conflito = supabase.table("convocacoes").select("*").gte("data", desde_conflitos).execute().data or []
        except Exception:
            regs_conflito = []
        conflitos_pendentes = []
        for reg in regs_conflito:
            meta_conf = obter_metadata_operacional(reg.get("observacao") or "")
            pend_conf = [c for c in (meta_conf.get("conflitos_convocacao") or []) if not c.get("resolvido")]
            if pend_conf:
                conflitos_pendentes.append((reg, pend_conf))
        if conflitos_pendentes:
            st.markdown("### 🚨 Conflitos de convocação")
            st.error(f"Existem {sum(len(p) for _, p in conflitos_pendentes)} tentativa(s) de dois supervisores convocarem a mesma pessoa.")
            for reg, pend_conf in conflitos_pendentes:
                colab_conf = dict_colaboradores.get(reg.get("colaborador_id"), {})
                detalhes_conf = ", ".join(
                    f"{x.get('tentativa_por','N/A')} em {str(x.get('em',''))[:16].replace('T',' ')}"
                    for x in pend_conf
                )
                cc1, cc2 = st.columns([5, 1])
                with cc1:
                    st.markdown(
                        f"**{colab_conf.get('nome','-')}** • {reg.get('data','')} • "
                        f"originalmente com **{reg.get('engenheiro','-')}**"
                    )
                    st.caption(f"Nova(s) tentativa(s): {detalhes_conf}")
                with cc2:
                    if st.button("Resolver", key=f"resolver_conf_{reg.get('id')}", use_container_width=True):
                        if resolver_conflitos_convocacao(reg):
                            st.rerun()

        st.markdown("### Ações rápidas")
        q1, q2, q3 = st.columns(3)
        q1.button("✅ FECHAR APONTAMENTOS", type="primary", use_container_width=True, on_click=_ir_menu_admin, args=("✅ APONTAMENTO",))
        q2.button("📋 PLANEJAR AMANHÃ", use_container_width=True, on_click=_ir_menu_admin, args=("📋 CONVOCAÇÃO",))
        q3.button("📊 VER RELATÓRIOS", use_container_width=True, on_click=_ir_menu_admin, args=("📊 RELATÓRIOS",))

    # --- DASHBOARD / AUDITORIA ---
    elif menu_escolhido == "🎛️ DASHBOARD":
        render_dashboard_consulta("admin_dash_melhorias")

    # --- 2. CONVOCAÇÃO ---
    elif menu_escolhido == "📋 CONVOCAÇÃO":
        st.markdown("## 📋 CONVOCAÇÃO E GERENCIAMENTO DE EQUIPE")
        tab_nova_conv, tab_corrigir_conv = st.tabs(["➕ Nova Convocação", "✏️ Correção / Exclusão Administrativa"])

        with tab_nova_conv:
            if obras:
                col_eng, col_info, col_turno = st.columns(3)
                with col_eng:
                    engenheiro_conv = st.selectbox("Engenheiro responsável:", ENGENHEIROS, key="eng_conv_adm")

                data_conv_auto = proximo_dia_util(datetime.date.today())
                with col_info:
                    st.info(f"📅 **Próximo dia:** {data_conv_auto.strftime('%d/%m/%Y')}")
                with col_turno:
                    turno_conv_adm = st.selectbox("Turno:", ["Integral", "Manhã", "Tarde", "Noite"], key="turno_conv_adm")

                unidades_unicas = UNIDADES_APROAR.copy()
                unidade_selecionada = st.selectbox("Unidade:", unidades_unicas, key="u_adm_sel")

                funcoes_disponiveis = sorted(list(set([c.get('funcao', '') for c in colaboradores if c.get('funcao')])))
                filtro_funcao_adm = st.selectbox("Filtrar por Função (Opcional):", ["TODAS"] + funcoes_disponiveis, key="f_adm_sel")

                if filtro_funcao_adm != "TODAS":
                    colabs_filtrados_adm = [c for c in colaboradores if c.get('funcao') == filtro_funcao_adm]
                else:
                    colabs_filtrados_adm = colaboradores

                mapa_colab_adm = {f"{c['nome']}  ({c.get('funcao','-')})": c['id'] for c in colabs_filtrados_adm}
                equipe_selecionada = st.multiselect(
                    "Buscar ou Selecionar Colaboradores Cadastrados:",
                    list(mapa_colab_adm.keys()),
                    key="eq_adm_sel"
                )

                st.markdown("#### ➕ Incluir nome digitado")
                avulso_adm = st.checkbox("É avulso?", key="avulso_conv_adm")
                nome_manual_adm = st.text_input(
                    "Nome do avulso:" if avulso_adm else "Adicionar colaborador pelo nome (opcional):",
                    placeholder="Digite o nome completo...",
                    key="nome_manual_conv_adm"
                )

                tipo_manual_adm = "Profissional"
                funcao_manual_adm = ""
                if avulso_adm or nome_manual_adm.strip():
                    ca1, ca2 = st.columns(2)
                    with ca1:
                        tipo_manual_adm = st.selectbox(
                            "Categoria da diária:",
                            ["Profissional", "Ajudante"],
                            key="tipo_manual_conv_adm"
                        )
                    with ca2:
                        funcao_manual_adm = st.text_input(
                            "Função (opcional):",
                            placeholder="Ex.: pintor, eletricista...",
                            key="funcao_manual_conv_adm"
                        )
                    st.caption(
                        f"Diária aplicada: Profissional = {formatar_reais(VALOR_DIARIA_PROFISSIONAL)} • "
                        f"Ajudante = {formatar_reais(VALOR_DIARIA_AJUDANTE)}"
                    )

                with st.container(border=True):
                    st.markdown(f"**Panorama de {engenheiro_conv} ({data_conv_auto.strftime('%d/%m/%Y')} • {turno_conv_adm})**")
                    try:
                        convs_eng_data = supabase.table("convocacoes").select("*").eq("engenheiro", engenheiro_conv).eq("data", data_conv_auto.isoformat()).execute().data
                    except Exception:
                        convs_eng_data = []
                    ids_ja_alocados_eng = {c['colaborador_id'] for c in convs_eng_data}
                    nomes_ja_alocados = [dict_colaboradores.get(cid, {}).get('nome', '') for cid in ids_ja_alocados_eng]
                    if nomes_ja_alocados:
                        st.caption("Já escalados por este engenheiro nesta data: " + ", ".join([n for n in nomes_ja_alocados if n]))
                    else:
                        st.caption("Nenhum escalado por este engenheiro ainda para o próximo dia.")
                    st.caption("A demanda específica será escolhida individualmente no Apontamento Diário.")

                if st.button("CONFIRMAR CONVOCAÇÃO", type="primary", use_container_width=True, key="btn_confirm_conv_adm"):
                    if avulso_adm and not nome_manual_adm.strip():
                        st.warning("Para convocar como avulso, digite o nome do colaborador.")
                    elif not equipe_selecionada and not nome_manual_adm.strip():
                        st.warning("Selecione um colaborador cadastrado ou digite um nome.")
                    else:
                        obra_id_placeholder = obter_obra_placeholder_unidade(unidade_selecionada)
                        if not obra_id_placeholder:
                            st.error(f"Não foi possível preparar a Unidade **{unidade_selecionada}** para a convocação.")
                        else:
                            pessoas = []
                            for label_colab in equipe_selecionada:
                                c_id = mapa_colab_adm[label_colab]
                                nome_existente = dict_colaboradores.get(c_id, {}).get('nome', label_colab.split('  (')[0])
                                pessoas.append((c_id, nome_existente))

                            if nome_manual_adm.strip():
                                c_id_manual, colab_manual, msg_manual = criar_ou_obter_colaborador_manual(
                                    nome_manual_adm,
                                    tipo_manual_adm,
                                    funcao_manual_adm,
                                    avulso=avulso_adm
                                )
                                if c_id_manual:
                                    pessoas.append((c_id_manual, colab_manual.get('nome', nome_manual_adm)))
                                    if "existente" in msg_manual.lower():
                                        st.info(msg_manual)
                                else:
                                    st.error(msg_manual)

                            pessoas_unicas = []
                            ids_vistos = set()
                            for cid, nome_pessoa in pessoas:
                                if cid and cid not in ids_vistos:
                                    pessoas_unicas.append((cid, nome_pessoa))
                                    ids_vistos.add(cid)

                            sucessos = 0
                            avisos = []
                            for c_id, nome_pessoa in pessoas_unicas:
                                ok, motivo = inserir_convocacao_segura(
                                    obra_id_placeholder,
                                    c_id,
                                    data_conv_auto,
                                    engenheiro_conv,
                                    turno_conv_adm
                                )
                                if ok:
                                    sucessos += 1
                                else:
                                    avisos.append(f"{nome_pessoa}: {motivo}.")

                            if sucessos:
                                limpar_cache_operacional()
                                st.success(
                                    f"✅ {sucessos} colaborador(es) convocado(s) para {unidade_selecionada} "
                                    f"em {data_conv_auto.strftime('%d/%m/%Y')} • Turno {turno_conv_adm}."
                                )
                            for aviso in avisos:
                                st.warning(aviso)
            else:
                st.info("Cadastre pelo menos uma Unidade/Obra na aba Configurações.")

        with tab_corrigir_conv:
            st.markdown("### ✏️ Correção / Exclusão Administrativa")
            st.write("Realocar um colaborador para outra Unidade/Obra ou excluir uma convocação lançada incorretamente pelo campo.")

            c_corr1, c_corr2 = st.columns(2)
            with c_corr1:
                data_corr = st.date_input(
                    "Data da Convocação para Corrigir:",
                    value=proximo_dia_util(datetime.date.today()),
                    format="DD/MM/YYYY",
                    key="d_corr"
                )

            try:
                convs_existentes = supabase.table("convocacoes").select("*").eq("data", data_corr.isoformat()).execute().data
            except Exception:
                convs_existentes = []

            if not convs_existentes:
                st.info("Nenhuma convocação encontrada nesta data para correção.")
            else:
                mapa_convs_corr = {}
                for item in convs_existentes:
                    colab_inf = dict_colaboradores.get(item['colaborador_id'], {})
                    obra_inf = dict_obras.get(item['obra_id'], {})
                    nome_obra_exib = obra_inf.get('nome', 'N/A')
                    if eh_obra_placeholder(obra_inf):
                        nome_obra_exib = "Obra/Serviço a definir"
                    rotulo = (
                        f"{colab_inf.get('nome','N/A')} | "
                        f"{obra_inf.get('unidade','N/A')} | {nome_obra_exib} | "
                        f"Eng: {item.get('engenheiro','N/A')}"
                    )
                    mapa_convs_corr[rotulo] = item

                with c_corr2:
                    conv_selecionada_rotulo = st.selectbox(
                        "Selecione o colaborador/convocação:",
                        list(mapa_convs_corr.keys()),
                        key="conv_sel_corr"
                    )

                registro_corr = mapa_convs_corr[conv_selecionada_rotulo]
                colab_corr = dict_colaboradores.get(registro_corr.get('colaborador_id'), {})
                obra_corr = dict_obras.get(registro_corr.get('obra_id'), {})
                unidade_atual_corr = obra_corr.get('unidade', '')
                nome_obra_atual_corr = obra_corr.get('nome', '')

                st.markdown(f"**Colaborador:** {colab_corr.get('nome', 'N/A')}")

                unidades_disponiveis = UNIDADES_APROAR.copy()
                idx_unidade = unidades_disponiveis.index(unidade_atual_corr) if unidade_atual_corr in unidades_disponiveis else 0

                c_dest1, c_dest2, c_dest3 = st.columns(3)
                with c_dest1:
                    nova_unidade_corr = st.selectbox(
                        "Unidade de destino:",
                        unidades_disponiveis,
                        index=idx_unidade,
                        key="u_dest_corr"
                    )

                obras_nova_u_lista = obras_reais_da_unidade(nova_unidade_corr)
                mapa_obras_nova_u = {o['nome']: o['id'] for o in obras_nova_u_lista}
                opcao_a_definir = "A DEFINIR NO APONTAMENTO"
                opcoes_obra_corr = [opcao_a_definir] + list(mapa_obras_nova_u.keys())

                if nova_unidade_corr == unidade_atual_corr and nome_obra_atual_corr in mapa_obras_nova_u:
                    idx_obra_corr = opcoes_obra_corr.index(nome_obra_atual_corr)
                else:
                    idx_obra_corr = 0

                with c_dest2:
                    nova_obra_corr = st.selectbox(
                        "Obra / Serviço de destino:",
                        opcoes_obra_corr,
                        index=idx_obra_corr,
                        key="o_dest_corr"
                    )
                with c_dest3:
                    eng_atual = registro_corr.get('engenheiro', ENGENHEIROS[0])
                    idx_eng = ENGENHEIROS.index(eng_atual) if eng_atual in ENGENHEIROS else 0
                    novo_eng_corr = st.selectbox(
                        "Engenheiro responsável:",
                        ENGENHEIROS,
                        index=idx_eng,
                        key="eng_dest_corr"
                    )

                b_corr1, b_corr2 = st.columns(2)
                with b_corr1:
                    if st.button("💾 SALVAR REALOCAÇÃO", type="primary", use_container_width=True):
                        if nova_obra_corr == opcao_a_definir:
                            nova_obra_id = obter_obra_placeholder_unidade(nova_unidade_corr)
                        else:
                            nova_obra_id = mapa_obras_nova_u.get(nova_obra_corr)

                        if not nova_obra_id:
                            st.error("Não foi possível definir a Unidade/Obra de destino.")
                        else:
                            supabase.table("convocacoes").update({
                                "obra_id": nova_obra_id,
                                "engenheiro": novo_eng_corr
                            }).eq("id", registro_corr['id']).execute()
                            st.success("✅ Colaborador realocado com sucesso.")
                            st.rerun()

                with b_corr2:
                    id_corr_atual = registro_corr["id"]

                    if st.button(
                        "🗑️ EXCLUIR CONVOCAÇÃO",
                        use_container_width=True,
                        key=f"exc_conv_{id_corr_atual}"
                    ):
                        st.session_state["conv_exclusao_pendente"] = id_corr_atual
                        st.rerun()

                    if st.session_state.get("conv_exclusao_pendente") == id_corr_atual:
                        st.markdown(
                            f"""
                            <div style="
                                background:#F8FAFC;
                                border:1px solid #CBD5E1;
                                border-left:4px solid #EF4444;
                                border-radius:12px;
                                padding:14px 16px;
                                margin:8px 0 12px 0;
                                color:#0F172A;
                                line-height:1.45;
                            ">
                                <div style="font-weight:700; font-size:15px; margin-bottom:4px;">
                                    Confirmar exclusão
                                </div>
                                <div style="font-size:14px;">
                                    Excluir a convocação de
                                    <strong>{colab_corr.get('nome', 'N/A')}</strong>
                                    em <strong>{data_corr.strftime('%d/%m/%Y')}</strong>?
                                </div>
                                <div style="font-size:12px; color:#64748B; margin-top:5px;">
                                    Esta ação não pode ser desfeita.
                                </div>
                            </div>
                            """,
                            unsafe_allow_html=True
                        )

                        conf1, conf2 = st.columns(2)

                        with conf1:
                            if st.button(
                                "✅ SIM, EXCLUIR",
                                type="primary",
                                use_container_width=True,
                                key=f"confirma_exc_conv_{id_corr_atual}"
                            ):
                                try:
                                    retorno_exc = (
                                        supabase.table("convocacoes")
                                        .delete()
                                        .eq("id", id_corr_atual)
                                        .execute()
                                    )

                                    st.session_state.pop("conv_exclusao_pendente", None)
                                    st.session_state["msg_corr_admin"] = (
                                        f"✅ Convocação de {colab_corr.get('nome', 'N/A')} excluída com sucesso."
                                    )
                                    limpar_cache_operacional()
                                    st.rerun()

                                except Exception as e:
                                    st.error(f"Não foi possível excluir a convocação: {e}")

                        with conf2:
                            if st.button(
                                "↩️ CANCELAR",
                                use_container_width=True,
                                key=f"cancela_exc_conv_{id_corr_atual}"
                            ):
                                st.session_state.pop("conv_exclusao_pendente", None)
                                st.rerun()

    # --- MENSAGEM PARA WHATSAPP ---
    elif menu_escolhido == "💬 WHATSAPP":
        st.markdown("## 💬 MENSAGEM DE CONVOCAÇÃO PARA WHATSAPP")
        st.write("Gere a divisão de equipes no padrão do grupo de Colaboradores e copie a mensagem pronta.")

        c_wpp1, c_wpp2 = st.columns([1, 1])
        with c_wpp1:
            data_wpp = st.date_input(
                "Data da divisão:",
                value=proximo_dia_util(datetime.date.today()),
                format="DD/MM/YYYY",
                key="data_mensagem_wpp"
            )
        with c_wpp2:
            mostrar_funcao_wpp = st.checkbox(
                "Mostrar função entre parênteses",
                value=False,
                key="mostrar_funcao_wpp"
            )

        try:
            convocacoes_wpp = (
                supabase.table("convocacoes")
                .select("*")
                .eq("data", data_wpp.isoformat())
                .execute().data or []
            )
        except Exception as e:
            convocacoes_wpp = []
            st.error(f"Não foi possível carregar as convocações: {e}")

        agrupado_wpp = organizar_convocacoes_whatsapp(convocacoes_wpp, mostrar_funcao=mostrar_funcao_wpp)
        unidades_com_divisao = list(agrupado_wpp.keys())

        unidades_cadastradas_wpp = sorted({
            o.get("unidade") for o in obras
            if o.get("unidade") and normalizar(o.get("unidade")) not in ["GERAL", "NAO IDENTIFICADA"]
        }, key=lambda x: normalizar(x))
        unidades_sem_divisao = [u for u in unidades_cadastradas_wpp if u not in unidades_com_divisao]

        if unidades_com_divisao:
            st.success(
                f"{len(convocacoes_wpp)} convocação(ões) encontrada(s) em "
                f"{len(unidades_com_divisao)} unidade(s)."
            )
        else:
            st.warning("Ainda não há nenhuma convocação registrada para a data selecionada.")

        if unidades_sem_divisao:
            with st.expander("🔎 Unidades sem convocação registrada nesta data"):
                st.write(", ".join(formatar_unidade_whatsapp(u) for u in unidades_sem_divisao))
                st.caption("Essa lista é apenas uma referência; podem existir unidades sem atividade nesta data.")

        aviso_pendentes_wpp = st.checkbox(
            "Ainda existem demandas que serão enviadas por outros responsáveis",
            value=bool(unidades_sem_divisao),
            key="aviso_pendentes_wpp",
            help="Ao marcar, o texto acrescenta: 'As demais demandas serão enviadas pelos respectivos responsáveis.'"
        )

        mensagem_geral_wpp = montar_mensagem_whatsapp(
            data_wpp,
            convocacoes_wpp,
            mostrar_funcao=mostrar_funcao_wpp,
            aviso_pendentes=aviso_pendentes_wpp
        )

        st.markdown("### 📋 Mensagem completa")
        st.caption("Use o ícone de copiar no canto do bloco abaixo e cole diretamente no WhatsApp.")
        st.code(mensagem_geral_wpp, language=None, wrap_lines=True)

        if unidades_com_divisao:
            st.markdown("### 🏢 Mensagem separada por Unidade")
            st.caption("Caso prefira enviar a divisão de cada Unidade separadamente.")
            for unidade in unidades_com_divisao:
                with st.expander(f"📌 {formatar_unidade_whatsapp(unidade)}"):
                    mensagem_unidade = montar_mensagem_whatsapp(
                        data_wpp,
                        convocacoes_wpp,
                        mostrar_funcao=mostrar_funcao_wpp,
                        aviso_pendentes=False,
                        somente_unidade=unidade
                    )
                    st.code(mensagem_unidade, language=None, wrap_lines=True)

        st.markdown("---")
        st.caption(
            "A mensagem usa somente a Unidade da convocação. Obra/Serviço não é exibida, "
            "e os colaboradores são numerados automaticamente."
        )

    # --- 3. APONTAMENTO ---
    elif menu_escolhido == "✅ APONTAMENTO":
        render_apontamento_operacional(None, "admin_apont_melhorias")

    # --- 4. RELATÓRIOS ---
    elif menu_escolhido == "📊 RELATÓRIOS":
        st.markdown("## 📊 RELATÓRIO DE CUSTOS E FECHAMENTO")
        
        # Filtro de Periodicidade
        c_p1, c_p2, c_p3 = st.columns(3)
        with c_p1:
            periodicidade = st.selectbox("Periodicidade:", ["Personalizado", "Diário", "Semanal", "Mensal"], key="rel_periodo")
        
        hoje = datetime.date.today()
        if periodicidade == "Diário":
            dt_inicio_def = hoje
            dt_fim_def = hoje
        elif periodicidade == "Semanal":
            dt_inicio_def = hoje - datetime.timedelta(days=7)
            dt_fim_def = hoje
        elif periodicidade == "Mensal":
            dt_inicio_def = hoje.replace(day=1)
            dt_fim_def = hoje
        else:
            dt_inicio_def = hoje
            dt_fim_def = hoje

        with c_p2:
            data_inicio_rel = st.date_input("Início:", value=dt_inicio_def, format="DD/MM/YYYY", key="data_ini_rel")
        with c_p3:
            data_fim_rel = st.date_input("Fim:", value=dt_fim_def, format="DD/MM/YYYY", key="data_fim_rel")

        col_r1, col_r2 = st.columns(2)
        with col_r1:
            eng_relatorio = st.selectbox("Engenheiro:", ["TODOS OS ENGENHEIROS"] + ENGENHEIROS, key="eng_rel")
        with col_r2:
            obras_rel_lista = sorted(list(set([o['nome'] for o in obras]))) if obras else []
            obra_relatorio = st.selectbox("Filtro por Obra:", ["TODAS AS OBRAS"] + obras_rel_lista, key="obra_rel")

        query_rel = supabase.table("convocacoes").select("*").gte("data", data_inicio_rel.isoformat()).lte("data", data_fim_rel.isoformat())
        if eng_relatorio != "TODOS OS ENGENHEIROS":
            query_rel = query_rel.eq("engenheiro", eng_relatorio)
        obra_id_filtro = None
        if obra_relatorio != "TODAS AS OBRAS":
            obra_id_filtro = next((o['id'] for o in obras if o['nome'] == obra_relatorio), None)

        dados_relatorio = query_rel.execute().data if data_inicio_rel <= data_fim_rel else []
        if obra_relatorio != "TODAS AS OBRAS" and obra_id_filtro:
            dados_relatorio = [
                row for row in dados_relatorio
                if str(row.get("obra_id")) == str(obra_id_filtro)
                or any(
                    normalizar(x.get("servico")) == normalizar(obra_relatorio)
                    for x in _normalizar_servicos_adicionais(obter_metadata_operacional(row.get("observacao") or ""))
                )
            ]

        col_btn1, col_btn2 = st.columns(2)

        with col_btn1:
            if st.button("📄 Gerar PDF", type="primary", use_container_width=True):
                try:
                    if data_inicio_rel > data_fim_rel:
                        st.error("Data inicial maior que a final.")
                    elif not dados_relatorio:
                        st.warning("Sem dados no período.")
                    else:
                        agrupado_eng = {}
                        for row in dados_relatorio:
                            eng = row.get('engenheiro', 'NÃO IDENTIFICADO')
                            ob = row['obra_id']
                            if eng not in agrupado_eng: agrupado_eng[eng] = {}
                            if ob not in agrupado_eng[eng]: agrupado_eng[eng][ob] = []
                            agrupado_eng[eng][ob].append(row)

                        pdf = FPDF(orientation='L')
                        for eng, obras_eng in agrupado_eng.items():
                            pdf.add_page()
                            pdf.set_font("Arial", 'B', 14)
                            pdf.cell(0, 10, txt=to_latin(f"APROAR - RELATÓRIO DE CUSTOS | ENG: {eng}"), ln=True, align='C')
                            pdf.set_font("Arial", size=10)
                            pdf.cell(0, 8, txt=to_latin(f"Período: {data_inicio_rel.strftime('%d/%m/%Y')} a {data_fim_rel.strftime('%d/%m/%Y')} ({periodicidade})"), ln=True, align='C')
                            pdf.ln(5)
                            
                            for o_id, apontamentos in obras_eng.items():
                                dados_ob = dict_obras.get(o_id, {"nome": "N/A", "unidade": "N/A"})
                                pdf.set_font("Arial", 'B', 10)
                                pdf.set_fill_color(30, 41, 59)
                                pdf.set_text_color(255, 255, 255)
                                pdf.cell(0, 7, txt=to_latin(f"Unidade: {dados_ob['unidade']} | Obra: {dados_ob['nome']}"), ln=True, fill=True)
                                pdf.set_text_color(0, 0, 0)
                                
                                pdf.set_font("Arial", 'B', 9)
                                pdf.cell(25, 6, to_latin("Data"), border=1, align='C')
                                pdf.cell(65, 6, to_latin("Colaborador"), border=1)
                                pdf.cell(50, 6, to_latin("Função"), border=1)
                                pdf.cell(32, 6, to_latin("Status"), border=1, align='C')
                                pdf.cell(24, 6, to_latin("Diária"), border=1, align='C')
                                pdf.cell(24, 6, to_latin("Extra"), border=1, align='C')
                                pdf.cell(51, 6, to_latin("Obs"), border=1, ln=True)
                                
                                pdf.set_font("Arial", '', 8)
                                for row in apontamentos:
                                    colab = dict_colaboradores.get(row['colaborador_id'], {})
                                    nome = colab.get('nome', 'N/A')
                                    funcao = colab.get('funcao', 'N/A')
                                    status = normalizar_status_operacional(row.get('status', 'Presente (Integral)'))
                                    extra = float(row.get('valor_extra', 0) or 0) if status_eh_presenca(status) else 0.0
                                    obs_livre_pdf = decompor_observacao_operacional(row.get('observacao', ''))[1]
                                    servicos_pdf = descricao_servicos_convocacao(row, dados_ob)
                                    obs = f"{servicos_pdf} | {obs_livre_pdf}" if obs_livre_pdf else servicos_pdf
                                    diaria_base = calcular_diaria_proporcional(status, obter_valor_diaria_colaborador(colab))
                                    
                                    pdf.cell(25, 6, to_latin(row.get('data', '')), border=1, align='C')
                                    pdf.cell(65, 6, to_latin(nome[:28]), border=1)
                                    pdf.cell(50, 6, to_latin(funcao[:20]), border=1)
                                    pdf.cell(32, 6, to_latin(status[:14]), border=1, align='C')
                                    pdf.cell(24, 6, to_latin(f"R$ {diaria_base:.2f}"), border=1, align='C')
                                    pdf.cell(24, 6, to_latin(f"R$ {extra:.2f}"), border=1, align='C')
                                    pdf.cell(51, 6, to_latin(obs[:30]), border=1, ln=True)
                                pdf.ln(3)

                        pdf_output = pdf.output(dest='S').encode('latin1')
                        st.download_button(
                            label="📥 Baixar PDF Gerado",
                            data=pdf_output,
                            file_name=f"relatorio_custos_{data_inicio_rel.strftime('%d-%m-%Y')}_a_{data_fim_rel.strftime('%d-%m-%Y')}.pdf",
                            mime="application/pdf"
                        )
                except Exception as e:
                    st.error(f"Erro ao gerar PDF: {e}")

        with col_btn2:
            if st.button("📊 Gerar Excel (Abas por Dia + Cores por Engenheiro)", use_container_width=True):
                try:
                    if data_inicio_rel > data_fim_rel:
                        st.error("Data inicial maior que a final.")
                    elif not dados_relatorio:
                        st.warning("Sem dados no período.")
                    else:
                        lista_excel = []
                        for row in dados_relatorio:
                            ob = dict_obras.get(row['obra_id'], {"nome": "N/A", "unidade": "N/A"})
                            colab = dict_colaboradores.get(row['colaborador_id'], {})
                            status = normalizar_status_operacional(row.get('status', 'Presente (Integral)'))
                            diaria_calc = float(calcular_diaria_proporcional(status, obter_valor_diaria_colaborador(colab)))
                            extra = float(row.get('valor_extra') or 0.0) if status_eh_presenca(status) else 0.0
                            
                            lista_excel.append({
                                "Data": str(row.get('data')),
                                "Engenheiro": str(row.get('engenheiro', 'N/A')),
                                "Unidade": str(ob['unidade']),
                                "Obra": str(descricao_servicos_convocacao(row, ob)),
                                "Colaborador": str(colab.get('nome', 'N/A')),
                                "Funcao": str(colab.get('funcao', 'N/A')),
                                "Status": str(status),
                                "Diaria": diaria_calc,
                                "Extra": extra,
                                "Observacao": str(decompor_observacao_operacional(row.get('observacao', ''))[1])
                            })
                        
                        df_excel = pd.DataFrame(lista_excel)
                        cores_engenheiros = {
                            "VICTOR": "E0F2FE", "EDUARDO": "DCFCE7", "GUSTAVO": "FEF9C3",
                            "JOEL": "F3E8FF", "NETO": "FFEDD5", "SOARES": "FFE4E6",
                            "GABRIEL": "CCFBF1", "PAULO": "F1F5F9"
                        }
                        
                        wb = openpyxl.Workbook()
                        wb.remove(wb.active)
                        
                        font_titulo = Font(name="Arial", size=11, bold=True, color="FFFFFF")
                        fill_cabecalho = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
                        font_obra_hdr = Font(name="Arial", size=10, bold=True, color="1E293B")
                        fill_obra_hdr = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
                        borda_fina = Border(
                            left=Side(style='thin', color='CBD5E1'), right=Side(style='thin', color='CBD5E1'),
                            top=Side(style='thin', color='CBD5E1'), bottom=Side(style='thin', color='CBD5E1')
                        )
                        
                        for data_str in sorted(df_excel['Data'].unique()):
                            df_dia = df_excel[df_excel['Data'] == data_str]
                            ws = wb.create_sheet(title=str(data_str))
                            
                            current_row = 1
                            ws.cell(row=current_row, column=1, value=f"APONTAMENTO DIÁRIO DE EQUIPES - DATA: {data_str}").font = Font(name="Arial", size=12, bold=True)
                            current_row += 2
                            
                            for unidade_nome in sorted(df_dia['Unidade'].unique()):
                                df_unidade = df_dia[df_dia['Unidade'] == unidade_nome]
                                for obra_nome in sorted(df_unidade['Obra'].unique()):
                                    df_obra = df_unidade[df_unidade['Obra'] == obra_nome]
                                    
                                    ws.cell(row=current_row, column=1, value=f"UNIDADE: {unidade_nome}  |  OBRA: {obra_nome}").font = font_obra_hdr
                                    for c_idx in range(1, 9):
                                        ws.cell(row=current_row, column=c_idx).fill = fill_obra_hdr
                                    current_row += 1
                                    
                                    colunas_tabela = ["Colaborador", "Função", "Engenheiro Resp.", "Status", "Diária (R$)", "Extra (R$)", "Custo Total (R$)", "Observação"]
                                    for c_idx, col_nome in enumerate(colunas_tabela, 1):
                                        cell = ws.cell(row=current_row, column=c_idx, value=col_nome)
                                        cell.font = font_titulo
                                        cell.fill = fill_cabecalho
                                        cell.alignment = Alignment(horizontal="center", vertical="center")
                                    current_row += 1
                                    
                                    inicio_dados_obra = current_row
                                    for _, r in df_obra.iterrows():
                                        eng_resp = r["Engenheiro"]
                                        cor_hex = cores_engenheiros.get(str(eng_resp).upper(), "FFFFFF")
                                        fill_engenheiro = PatternFill(start_color=cor_hex, end_color=cor_hex, fill_type="solid")
                                        
                                        celula_custo_formula = f"=E{current_row}+F{current_row}"
                                        linha_dados = [
                                            r["Colaborador"], r["Funcao"], r["Engenheiro"], r["Status"],
                                            r["Diaria"], r["Extra"], celula_custo_formula, r["Observacao"]
                                        ]
                                        
                                        for c_idx, val in enumerate(linha_dados, 1):
                                            c_cell = ws.cell(row=current_row, column=c_idx, value=val)
                                            c_cell.font = Font(name="Arial", size=9)
                                            c_cell.border = borda_fina
                                            c_cell.fill = fill_engenheiro 
                                            
                                            if c_idx in [5, 6, 7]:
                                                c_cell.number_format = 'R$ #,##0.00'
                                                c_cell.alignment = Alignment(horizontal="right")
                                            elif c_idx in [3, 4]:
                                                c_cell.alignment = Alignment(horizontal="center")
                                        current_row += 1
                                    
                                    fim_dados_obra = current_row - 1
                                    ws.cell(row=current_row, column=5, value=f"TOTAL OBRA {obra_nome}:").font = Font(name="Arial", size=10, bold=True)
                                    ws.cell(row=current_row, column=5).alignment = Alignment(horizontal="right")
                                    
                                    celula_subtotal = ws.cell(row=current_row, column=7, value=f"=SUM(G{inicio_dados_obra}:G{fim_dados_obra})")
                                    celula_subtotal.font = Font(name="Arial", size=10, bold=True)
                                    celula_subtotal.number_format = 'R$ #,##0.00'
                                    celula_subtotal.border = borda_fina
                                    current_row += 2

                            for col in ws.columns:
                                max_len = 0
                                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                                for cell in col:
                                    if cell.row in [1, 2, 3] or (cell.value and str(cell.value).startswith("UNIDADE:")):
                                        continue
                                    if cell.value:
                                        val_str = str(cell.value)
                                        if len(val_str) > max_len:
                                            max_len = len(val_str)
                                ws.column_dimensions[col_letter].width = max(min(max_len + 4, 35), 14)

                        buffer = io.BytesIO()
                        wb.save(buffer)
                        
                        st.download_button(
                            label="📥 Baixar Excel (Abas por Dia + Cores de Engenheiros)",
                            data=buffer.getvalue(),
                            file_name=f"apontamentos_por_dia_{data_inicio_rel.strftime('%d-%m-%Y')}_a_{data_fim_rel.strftime('%d-%m-%Y')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                except Exception as e:
                    st.error(f"Erro ao gerar Excel: {e}")

    # --- 5. INDICADORES ---
    elif menu_escolhido == "📈 INDICADORES":
        render_indicadores_cumprimento("admin_ind_melhorias", None, mostrar_absenteismo=True)

    # --- INDISPONIBILIDADE ---
    elif menu_escolhido == "🚫 INDISPONIBILIDADE":
        render_indisponibilidades_admin()

    # --- 6. DISPONIBILIDADE ---
    elif menu_escolhido == "👥 DISPONIBILIDADE":
        render_aba_disponibilidade("admin")

    # --- 7. CONFIGURAÇÕES E SINCRONIZAÇÃO TRELLO ---
    elif menu_escolhido == "⚙️ CONFIGURAÇÕES":
        st.markdown("## ⚙️ CONFIGURAÇÕES E GERENCIAMENTO")
        
        # Sincronização Dinâmica Trello (mês vigente, lista manual ou busca de card/lista)
        with st.container(border=True):
            st.markdown("### 🔄 Sincronização com o Trello")
            st.write("Sincronize o mês vigente ou localize manualmente listas e cards de medições anteriores.")

            lists_trello, cards_trello = obter_listas_trello()
            mapa_nome_lista = {l.get('id'): l.get('name', 'Lista sem nome') for l in lists_trello}

            if st.session_state.get("trello_usando_snapshot"):
                st.info(
                    "ℹ️ O Trello está demorando a responder. "
                    "Estou usando a última leitura válida salva para manter o sistema operacional."
                )

            c_tr1, c_tr2 = st.columns(2)
            with c_tr1:
                if st.button("🚀 SINCRONIZAR MÊS VIGENTE (AUTOMÁTICO)", type="primary"):
                    with st.spinner("Sincronizando mês vigente..."):
                        sucesso, me = executar_sincronizacao_trello(listas_precarregadas=lists_trello, cards_precarregados=cards_trello)
                        if sucesso:
                            st.success(me)
                            st.rerun()
                        else:
                            st.error(me)

            with c_tr2:
                listas_abertas = [l for l in lists_trello if not l.get('closed', False)]
                if listas_abertas:
                    mapa_listas = {l['name']: l['id'] for l in listas_abertas}
                    lista_manual_sel = st.selectbox("Selecionar lista diretamente:", list(mapa_listas.keys()), key="sel_trello_manual")
                    if st.button("🔄 SINCRONIZAR LISTA SELECIONADA"):
                        id_sel = mapa_listas[lista_manual_sel]
                        with st.spinner(f"Sincronizando {lista_manual_sel}..."):
                            sucesso, me = executar_sincronizacao_trello(id_lista_target=id_sel, listas_precarregadas=lists_trello, cards_precarregados=cards_trello)
                            if sucesso:
                                st.success(me)
                                st.rerun()
                            else:
                                st.error(me)
                else:
                    st.warning(
                        "Trello temporariamente indisponível e ainda não há uma leitura anterior salva. "
                        "As obras já cadastradas continuam disponíveis normalmente."
                    )
                    detalhe_trello = st.session_state.get("trello_ultimo_erro", "")
                    st.caption("Quadro público • sem API Key ou Token.")
                    if detalhe_trello:
                        with st.expander("Diagnóstico técnico"):
                            st.code(detalhe_trello)

            st.markdown("#### 🔎 Busca manual para medições retroativas")
            termo_trello = st.text_input(
                "Buscar card ou lista por nome:",
                placeholder="Ex.: MEDIÇÃO JUNHO 2026, APRL005, MARACANAÚ...",
                key="busca_trello_retroativa"
            )

            if termo_trello.strip():
                termo_norm = normalizar(termo_trello)
                resultados = {}

                for lst in lists_trello:
                    if termo_norm in normalizar(lst.get('name', '')):
                        rotulo = f"📋 LISTA | {lst.get('name', 'Sem nome')}"
                        resultados[rotulo] = ("lista", lst.get('id'))

                for card in cards_trello:
                    if termo_norm in normalizar(card.get('name', '')):
                        nome_lista = mapa_nome_lista.get(card.get('idList'), 'Lista não identificada')
                        situacao = "arquivado" if card.get('closed', False) else "ativo"
                        rotulo = f"🗂️ CARD | {card.get('name', 'Sem nome')} | {nome_lista} | {situacao} | {str(card.get('id',''))[-6:]}"
                        resultados[rotulo] = ("card", card.get('id'))

                if resultados:
                    resultado_sel = st.selectbox("Resultados encontrados:", list(resultados.keys()), key="resultado_busca_trello")
                    tipo_resultado, id_resultado = resultados[resultado_sel]
                    if st.button("➕ SINCRONIZAR RESULTADO DA BUSCA", type="primary", use_container_width=True):
                        with st.spinner("Sincronizando resultado selecionado..."):
                            if tipo_resultado == "lista":
                                sucesso, me = executar_sincronizacao_trello(id_lista_target=id_resultado, listas_precarregadas=lists_trello, cards_precarregados=cards_trello)
                            else:
                                sucesso, me = executar_sincronizacao_trello(id_card_target=id_resultado, listas_precarregadas=lists_trello, cards_precarregados=cards_trello)
                            if sucesso:
                                st.success(me)
                                st.rerun()
                            else:
                                st.error(me)
                else:
                    st.info("Nenhum card ou lista encontrado para esse termo.")

        st.markdown("---")
        tab_cad_obra, tab_cad_colab, tab_import_colab, tab_limpeza = st.tabs(["🏗️ Obras", "👷 Colaboradores", "📤 Importar Colaboradores", "🗑️ Limpeza de Dados"])
        
        with tab_cad_obra:
            st.markdown("### Cadastrar Nova Obra")
            with st.form("form_cad_obra"):
                nome_obra = st.text_input("Nome da Obra (Ex: 1863, 1383...):")
                unidade_obra = st.text_input("Unidade (Ex: CENTRO, MUSEU, FIEC...):")
                submit_obra = st.form_submit_button("Cadastrar Obra")
                if submit_obra:
                    if nome_obra and unidade_obra:
                        supabase.table("obras").insert({"nome": nome_obra, "unidade": unidade_obra.upper()}).execute()
                        limpar_cache_operacional()
                        st.success("Obra cadastrada com sucesso!")
                        st.rerun()
                    else:
                        st.warning("Preencha todos os campos.")

        with tab_cad_colab:
            st.markdown("### Cadastrar Novo Colaborador")
            with st.form("form_cad_colab"):
                nome_colab = st.text_input("Nome Completo:")
                tipo_colab = st.selectbox("Categoria da diária:", ["Profissional", "Ajudante"], key="tipo_cad_colab")
                funcao_colab = st.text_input("Função / Cargo:")
                diaria_colab = valor_diaria_por_tipo(tipo_colab)
                st.caption(f"Valor aplicado automaticamente: {formatar_reais(diaria_colab)}")
                submit_colab = st.form_submit_button("Cadastrar Colaborador")
                if submit_colab:
                    if nome_colab:
                        funcao_salva = limpar_funcao(funcao_colab) if funcao_colab.strip() else tipo_colab.upper()
                        try:
                            supabase.table("colaboradores").insert({
                                "nome": nome_colab.strip().upper(),
                                "funcao": funcao_salva,
                                "valor_diaria": diaria_colab
                            }).execute()
                            limpar_cache_operacional()
                            st.success("Colaborador cadastrado com sucesso!")
                            st.rerun()
                        except Exception:
                            st.error("Não foi possível cadastrar. Verifique se esse nome já existe.")
                    else:
                        st.warning("Informe o nome do colaborador.")

        with tab_import_colab:
            st.markdown("### 📤 Importar Planilha de Colaboradores")
            st.write(
                "Envie uma planilha para adicionar novos colaboradores ou atualizar cadastros existentes. "
                "A conferência é feita pelo nome e nenhum colaborador ausente da planilha será excluído."
            )

            if st.session_state.get("msg_import_colab"):
                st.success(st.session_state.pop("msg_import_colab"))

            arquivo_import = st.file_uploader(
                "Selecione a planilha de colaboradores:",
                type=["xlsx", "xls", "csv"],
                key="arquivo_import_colaboradores"
            )

            if arquivo_import is not None:

                def _localizar_linha_cabecalho_colaboradores(df_bruto, limite=20):
                    """
                    Procura automaticamente a linha de cabeçalho da tabela.
                    Isso permite importar planilhas que tenham título, total, observações
                    ou linhas em branco antes de 'Código | Nome | ...'.
                    """
                    candidatos_nome = {
                        "NOME",
                        "NOME COMPLETO",
                        "COLABORADOR",
                        "FUNCIONARIO",
                        "FUNCIONÁRIO",
                        "EMPREGADO",
                    }

                    qtd = min(len(df_bruto), limite)

                    for idx in range(qtd):
                        valores = []

                        for valor in df_bruto.iloc[idx].tolist():
                            if pd.isna(valor):
                                continue

                            txt = normalizar(str(valor))
                            if txt:
                                valores.append(txt)

                        # Prioridade: linha que realmente contém uma coluna de nome.
                        if any(v in candidatos_nome for v in valores):
                            return idx

                        # Também aceita cabeçalhos descritivos como
                        # "Nome do colaborador", sem confundir o título
                        # "COLABORADORES ADMITIDOS / ATIVOS" com cabeçalho.
                        for v in valores:
                            tokens = [t for t in re.split(r"[^A-Z0-9]+", v) if t]
                            if "NOME" in tokens:
                                return idx

                    return None


                def _ler_planilha_colaboradores(arquivo):
                    """
                    Lê XLS/XLSX/CSV detectando automaticamente o cabeçalho.
                    Retorna (dataframe, linha_cabecalho_1_based).
                    """
                    nome_arquivo = arquivo.name.lower()

                    if nome_arquivo.endswith(".csv"):
                        arquivo.seek(0)

                        try:
                            bruto = pd.read_csv(
                                arquivo,
                                sep=None,
                                engine="python",
                                header=None
                            )
                            encoding_usado = None

                        except UnicodeDecodeError:
                            arquivo.seek(0)
                            bruto = pd.read_csv(
                                arquivo,
                                sep=None,
                                engine="python",
                                header=None,
                                encoding="latin-1"
                            )
                            encoding_usado = "latin-1"

                        linha_header = _localizar_linha_cabecalho_colaboradores(bruto)

                        arquivo.seek(0)

                        kwargs = {
                            "sep": None,
                            "engine": "python",
                            "header": linha_header if linha_header is not None else 0,
                        }

                        if encoding_usado:
                            kwargs["encoding"] = encoding_usado

                        df = pd.read_csv(
                            arquivo,
                            **kwargs
                        )

                    else:
                        arquivo.seek(0)

                        # Primeiro lê sem cabeçalho para encontrar onde a tabela começa.
                        bruto = pd.read_excel(
                            arquivo,
                            header=None
                        )

                        linha_header = _localizar_linha_cabecalho_colaboradores(bruto)

                        arquivo.seek(0)

                        # Se não localizar nada, mantém compatibilidade com planilhas
                        # convencionais cujo cabeçalho já está na primeira linha.
                        df = pd.read_excel(
                            arquivo,
                            header=linha_header if linha_header is not None else 0
                        )

                    # Remove linhas e colunas completamente vazias.
                    df = df.dropna(axis=0, how="all").dropna(axis=1, how="all")

                    return df, (
                        linha_header + 1
                        if linha_header is not None
                        else 1
                    )


                try:
                    df_import, linha_cabecalho_detectada = _ler_planilha_colaboradores(
                        arquivo_import
                    )

                except Exception as e:
                    df_import = pd.DataFrame()
                    linha_cabecalho_detectada = None
                    st.error(f"Não foi possível ler a planilha: {e}")


                if not df_import.empty:

                    if linha_cabecalho_detectada:
                        st.caption(
                            f"✅ Cabeçalho da tabela identificado automaticamente na linha "
                            f"{linha_cabecalho_detectada}."
                        )
                    df_import.columns = [str(c).strip() for c in df_import.columns]
                    colunas = list(df_import.columns)

                    def localizar_coluna_import(candidatos):
                        mapa = {normalizar(c): c for c in colunas}
                        for candidato in candidatos:
                            if normalizar(candidato) in mapa:
                                return mapa[normalizar(candidato)]
                        for c in colunas:
                            c_norm = normalizar(c)
                            if any(normalizar(cand) in c_norm for cand in candidatos):
                                return c
                        return None

                    col_nome_auto = localizar_coluna_import([
                        "NOME", "NOME COMPLETO", "COLABORADOR", "FUNCIONARIO", "FUNCIONÁRIO", "EMPREGADO"
                    ])
                    col_funcao_auto = localizar_coluna_import([
                        "FUNCAO", "FUNÇÃO", "CARGO", "FUNCAO/CARGO", "FUNÇÃO/CARGO"
                    ])
                    col_categoria_auto = localizar_coluna_import([
                        "CATEGORIA", "TIPO", "CLASSIFICACAO", "CLASSIFICAÇÃO"
                    ])
                    col_avulso_auto = localizar_coluna_import(["AVULSO"])

                    c_imp1, c_imp2 = st.columns(2)
                    with c_imp1:
                        idx_nome = colunas.index(col_nome_auto) if col_nome_auto in colunas else 0
                        col_nome_import = st.selectbox(
                            "Coluna do NOME:",
                            colunas,
                            index=idx_nome,
                            key="map_nome_import"
                        )
                    with c_imp2:
                        opcoes_funcao = ["(não usar)"] + colunas
                        idx_funcao = opcoes_funcao.index(col_funcao_auto) if col_funcao_auto in colunas else 0
                        col_funcao_import = st.selectbox(
                            "Coluna da FUNÇÃO/CARGO:",
                            opcoes_funcao,
                            index=idx_funcao,
                            key="map_funcao_import"
                        )

                    c_imp3, c_imp4 = st.columns(2)
                    with c_imp3:
                        opcoes_categoria = ["(inferir pela função)"] + colunas
                        idx_categoria = opcoes_categoria.index(col_categoria_auto) if col_categoria_auto in colunas else 0
                        col_categoria_import = st.selectbox(
                            "Coluna PROFISSIONAL/AJUDANTE (opcional):",
                            opcoes_categoria,
                            index=idx_categoria,
                            key="map_categoria_import"
                        )
                    with c_imp4:
                        opcoes_avulso = ["(não usar)"] + colunas
                        idx_avulso = opcoes_avulso.index(col_avulso_auto) if col_avulso_auto in colunas else 0
                        col_avulso_import = st.selectbox(
                            "Coluna AVULSO (opcional):",
                            opcoes_avulso,
                            index=idx_avulso,
                            key="map_avulso_import"
                        )

                    registros_por_nome = {}
                    linhas_invalidas = 0
                    for _, linha in df_import.iterrows():
                        nome_val = linha.get(col_nome_import)
                        if pd.isna(nome_val) or not str(nome_val).strip():
                            linhas_invalidas += 1
                            continue

                        nome_limpo = " ".join(str(nome_val).strip().split()).upper()

                        funcao_val = ""
                        if col_funcao_import != "(não usar)":
                            bruto_funcao = linha.get(col_funcao_import)
                            if not pd.isna(bruto_funcao):
                                funcao_val = str(bruto_funcao).strip()

                        if col_categoria_import != "(inferir pela função)":
                            bruto_categoria = linha.get(col_categoria_import)
                            categoria_txt = "" if pd.isna(bruto_categoria) else normalizar(bruto_categoria)
                            if "AJUD" in categoria_txt or "AUX" in categoria_txt or "SERVENT" in categoria_txt:
                                tipo_val = "Ajudante"
                            elif "PROF" in categoria_txt:
                                tipo_val = "Profissional"
                            else:
                                tipo_val = inferir_tipo_colaborador(funcao_val)
                        else:
                            tipo_val = inferir_tipo_colaborador(funcao_val)

                        avulso_val = False
                        if col_avulso_import != "(não usar)":
                            bruto_avulso = linha.get(col_avulso_import)
                            if not pd.isna(bruto_avulso):
                                avulso_txt = normalizar(bruto_avulso)
                                avulso_val = avulso_txt in ["SIM", "S", "TRUE", "VERDADEIRO", "1", "X"]

                        if not funcao_val:
                            funcao_val = tipo_val.upper()
                        if avulso_val and not normalizar(funcao_val).startswith("AVULSO -"):
                            funcao_val = f"AVULSO - {funcao_val}"

                        funcao_salva = limpar_funcao(funcao_val)
                        diaria_salva = valor_diaria_por_tipo(tipo_val)

                        # Se o mesmo nome aparecer mais de uma vez na planilha, mantém a última ocorrência.
                        registros_por_nome[normalizar(nome_limpo)] = {
                            "nome": nome_limpo,
                            "funcao": funcao_salva,
                            "tipo": tipo_val,
                            "valor_diaria": diaria_salva,
                            "avulso": avulso_val,
                        }

                    registros_import = list(registros_por_nome.values())

                    if registros_import:
                        preview_import = pd.DataFrame([
                            {
                                "Nome": r["nome"],
                                "Função": r["funcao"],
                                "Categoria": r["tipo"],
                                "Diária": formatar_reais(r["valor_diaria"]),
                                "Avulso": "SIM" if r["avulso"] else "NÃO",
                            }
                            for r in registros_import
                        ])
                        st.caption(
                            f"{len(registros_import)} colaborador(es) pronto(s) para importar. "
                            + (f"{linhas_invalidas} linha(s) sem nome foram ignoradas." if linhas_invalidas else "")
                        )
                        st.dataframe(preview_import, use_container_width=True, hide_index=True)

                        if st.button(
                            "📤 IMPORTAR / ATUALIZAR COLABORADORES",
                            type="primary",
                            use_container_width=True,
                            key="btn_importar_colaboradores"
                        ):
                            try:
                                atuais_import = supabase.table("colaboradores").select("*").execute().data or []
                                mapa_atuais = {
                                    normalizar(c.get("nome", "")): c
                                    for c in atuais_import
                                    if c.get("nome")
                                }

                                novos = 0
                                atualizados = 0
                                erros = []

                                for reg in registros_import:
                                    existente = mapa_atuais.get(normalizar(reg["nome"]))
                                    payload = {
                                        "nome": reg["nome"],
                                        "funcao": reg["funcao"],
                                        "valor_diaria": reg["valor_diaria"],
                                    }
                                    try:
                                        if existente:
                                            supabase.table("colaboradores").update(payload).eq("id", existente["id"]).execute()
                                            atualizados += 1
                                        else:
                                            retorno = supabase.table("colaboradores").insert(payload).execute().data or []
                                            novos += 1
                                            if retorno:
                                                mapa_atuais[normalizar(reg["nome"])] = retorno[0]
                                    except Exception:
                                        erros.append(reg["nome"])

                                limpar_cache_operacional()
                                mensagem = f"Importação concluída: {novos} novo(s) e {atualizados} atualizado(s)."
                                if erros:
                                    mensagem += f" Não foi possível importar {len(erros)} registro(s)."
                                st.session_state["msg_import_colab"] = mensagem
                                st.rerun()
                            except Exception as e:
                                st.error(f"Erro durante a importação: {e}")
                    else:
                        st.warning("A planilha não possui colaboradores válidos para importar.")
                elif arquivo_import is not None:
                    st.warning("A planilha está vazia ou não pôde ser interpretada.")

        with tab_limpeza:
            st.markdown("### 🗑️ Limpeza e Manutenção de Registros")
            st.write("Use esta seção para remover registros incorretos ou limpar dados antigos de convocações/apontamentos.")
            
            data_limpeza = st.date_input("Selecionar data para limpeza de convocações:", value=datetime.date.today(), format="DD/MM/YYYY")
            if st.button("🗑️ EXCLUIR CONVOCAÇÕES DESTA DATA", type="primary"):
                try:
                    supabase.table("convocacoes").delete().eq("data", data_limpeza.isoformat()).execute()
                    st.success(f"Todas as convocações do dia {data_limpeza.strftime('%d/%m/%Y')} foram removidas com sucesso!")
                    st.rerun()
                except Exception as e:
                    st.error(f"Erro ao limpar dados: {e}")
